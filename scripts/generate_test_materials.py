#!/usr/bin/env python3
"""外注テスト資材(⑧)を生成する。

出力:
    outputs/外注テスト手順書_<日付>.xlsx … 非エンジニア向けの準備・進め方
    outputs/外注テスト項目書_<日付>.xlsx …
        - シート「A_機能テスト」: 仕様書v2.0の機能一覧・画面仕様から
          ランダムピックアップした検証項目(乱数シード固定で再現可能)
        - シート「B_要望確認テスト」: 2026-03-14以降に起票された要望の全件
          (No.132〜149＋6/3起票分)を1件1ケース化
"""
import os
import random
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TODAY = os.environ.get('BUILD_DATE') or date.today().isoformat()
SEED = 20260702           # ランダムピックアップの再現用シード
SAMPLE_SIZE = 15          # 機能テストの抽出数

HEADER_FILL = PatternFill('solid', fgColor='2563EB')
SECTION_FILL = PatternFill('solid', fgColor='DBEAFE')
BORDER = Border(*(Side(style='thin', color='CBD5E1'),) * 4)

# --- 仕様書v2.0由来の機能テスト候補プール(画面/手順/期待結果) ------------
POOL = [
    ('ログイン', 'テスト用IDと正しいパスワードを入力し「LOGIN」を押す',
     'ステージ選択画面（Select a Stage）に移動する'),
    ('ログイン', 'わざと間違ったパスワードを入力し「LOGIN」を押す',
     '赤い文字でエラーメッセージが表示され、画面は移動しない'),
    ('ログイン', '「LOGIN」を押した直後のボタンを見る',
     'ボタンが「Logging in...」に変わり、連打しても反応しない'),
    ('ステージ選択', '学年・パート・サブパートのプルダウンを開く',
     '自分が解放しているステージまでしか表示されない（先のステージは選べない）'),
    ('ステージ選択', '「Ranking 🏆」を押す', 'ランキング画面に移動する'),
    ('ゲーム', 'ステージを選んで「Game Start」→「Start」を押す',
     '最初に回答条件（Requirement）が表示され、Startでゲームが始まる'),
    ('ゲーム', '1ステージを最後までプレイし、出題数を数える',
     '問題は全部で8問（最初の1問はデモ）出題される'),
    ('ゲーム', '1問目（デモ）の動きを見る',
     '自動で正解が表示されて次に進む。タイマーは動かない'),
    ('ゲーム', '2問目以降で問題の音声を聞く',
     '問題文が2回読み上げられてから回答できるようになる'),
    ('ゲーム', '2問目以降でタイマーを見る',
     '30秒からカウントダウンし、残り10秒前後で赤色になる'),
    ('ゲーム', '何も答えずに時間切れまで待つ',
     '正解が音声と文字で表示され、自動で次の問題に進む（固まらない）'),
    ('ゲーム', 'マイクで正しい答えを言う',
     'ビーム→敵の爆発→緑の正解パネルが表示され、次の問題へ進む'),
    ('ゲーム', 'わざと違う答えを言う',
     '敵の攻撃演出の後、残り時間内であればもう一度答えられる'),
    ('ゲーム', '銃ボタンを押してマイクをONにする',
     '右上が「🎤 MIC: ON」になり、話した内容が「Heard: …」に表示される'),
    ('ゲーム', '「やめる」ボタンを押す', 'ステージ選択画面に戻る'),
    ('ゲーム', 'イラストつきの問題を確認する（例: 3-8-2）',
     '問題文（左）とイラスト（右）が重ならずに表示される'),
    ('結果', 'ステージを最後までプレイする',
     '正解数（○/7）とパーセントが表示される'),
    ('結果', '5問以上正解してクリアする',
     '「CLEAR」表示になり、次のステージが選べるようになる'),
    ('結果', '5問未満で終わる',
     '「あと N 問正解でクリア！」のNが正しい（例: 3問正解なら「あと2問」）'),
    ('結果', '同じステージを10回プレイする（クリアしなくてよい）',
     '10回目のあと、次のステージが自動で解放される'),
    ('ランキング', 'ランキング画面を開く',
     '今月の「Number of try」と「Best Scores」の上位3名が表示される'),
    ('管理画面', '（先生用）名前とニックネームを入れて「登録」を押す',
     '新しいユーザーIDとパスワードが画面に表示される'),
    ('管理画面', '（先生用）ユーザーの「編集」で学年・パートを変えて「保存」',
     '一覧の表示が変わり、その生徒は変更後のステージから遊べる'),
    ('管理画面', '（先生用）ユーザーIDを入れて「パスワードリセット」',
     '新しいパスワードが表示され、そのパスワードでログインできる'),
    ('全体', 'iPadまたはiPhoneを横向きにしてプレイする',
     '文字やボタンがはみ出したり重なったりしない'),
    ('全体', 'ゲーム中に画面が固まった場合の動きを見る（発生した場合のみ）',
     '20秒ほどで「画面が停止しました」の画面が出て「リトライ/次の問題へ/やめる」を選べる'),
]

# --- 2026-03-14以降の要望確認テスト(No, 内容要約, 画面, 手順, 期待結果) ----
REQUESTS = [
    ('No.132', '1-24/1-25の音声が出ない', 'ゲーム',
     '中1のパート24-1、24-2、25-1をプレイする',
     '問題・正解の音声が正しく再生され、イラストも表示される'),
    ('No.133', '「あと何問でクリア」の数が誤っている', '結果',
     '7問中4問正解で終える', '「あと1問正解でクリア！」と表示される（5問でクリアのため）'),
    ('No.134', '時間切れ後に固まる', 'ゲーム',
     '制限時間ぎりぎりまで問題を解き、時間切れさせる',
     '固まらずに正解表示→次の問題へ進む'),
    ('No.135', 'ログインボタンの反応がわからない', 'ログイン',
     'LOGINボタンを押す', 'ボタンの表示が変わり、処理中であることがわかる'),
    ('No.136', 'ログインに1分かかる', 'ログイン',
     'しばらく（30分以上）誰も使っていない状態でログインする',
     '時間がかかる場合「サーバーを起動しています…」の案内が出る。案内どおり待つとログインできる'),
    ('No.137', '固まったときの復帰が遅い', 'ゲーム',
     '（発生した場合のみ）ゲームが固まったときの時間を計る',
     '20秒ほどで「リトライ/次の問題へ/やめる」の画面が出る'),
    ('No.138', 'readが過去形の発音になる', 'ゲーム',
     '1-39-2の2問目「I read this book.」の音声を聞く',
     '「リード」（現在形）と発音される'),
    ('No.139', '認証が必要ですの意味がわからない', '全体',
     '1日以上あけてから、開いたままの画面で操作する',
     '「セッションの有効期限が切れました。ログインし直してください。」と表示されログイン画面に移動する'),
    ('No.140', '翌日スタートを押すと承認が必要と言われ進めない', 'ステージ選択',
     '前日に開いたままの画面で「Game Start」を押す',
     '自動でログイン画面に移動し、ログインし直すと続きから遊べる'),
    ('No.141', 'イラストと問題文が重なる', 'ゲーム',
     '中2のイラストつき問題（例: 2-45-2）をプレイする',
     '問題文とイラストが重ならない'),
    ('No.142', '2-22-1-7/2-24-1-2のread発音', 'ゲーム',
     '該当の問題の音声を聞く', '「リード」（現在形）と発音される'),
    ('No.143', '問題文の最初の音が聞こえない', 'ゲーム',
     '2-47-2の1問目の音声を聞く', '文の最初から欠けずに聞こえる'),
    ('No.144', '3-8-2の追加とイラスト', 'ステージ選択/ゲーム',
     '中3のパート8でサブパート2を選んでプレイする',
     '3-8-2が選択でき、8問すべてにイラストが表示される'),
    ('No.145', '回答条件が1度しか読まれない', 'ゲーム',
     '3-18-1、3-19-1をプレイし、問題音声の回数を数える',
     '各問題の問題文が毎回2回読み上げられる（※画面の回答条件は音声では読まれません）'),
    ('No.146', '3-20-1-3の問題文が1回しか読まれない', 'ゲーム',
     '3-20-1の3問目の音声を聞く', '問題文が2回読み上げられる'),
    ('No.147', '3-23-2が7問までしか読まれない', 'ゲーム',
     '3-23-2を最後までプレイする', 'デモ+7問=8問すべて出題される'),
    ('No.148', '3-24以降がNo Data', 'ステージ選択/ゲーム',
     '中3のパート24〜35を選んでプレイする', 'すべてのパートで問題が表示される'),
    ('No.149', '3-34-2の答えが認識されづらい', 'ゲーム',
     "3-34-2で「If she were a cat, she'd sleep all day.」のように短縮形でも答えてみる",
     '短縮形でも正解になる（完全な言い方でも正解になる）'),
    ('6/3-1', '1-1-1が認識されない（特に8番）', 'ゲーム',
     '1-1-1で「I」「You」など1語の答えを言う。文で言ってしまってもよい',
     '答えの単語が入っていれば正解になる'),
    ('6/3-2', 'readがレッドと発音される', 'ゲーム',
     '1-46-2-2 / 2-38-2-2 / 2-41-2-8 / 2-42-1-7 / 2-47-1-7 の音声を聞く',
     'すべて「リード」（現在形）と発音される'),
    ('6/3-3', '（同上・対象一覧の確認）', 'ゲーム',
     '上記5問以外にreadを含む問題があれば音声を聞く', '「リード」と発音される'),
]


def style_header(ws, columns):
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 26
    for i, (name, width) in enumerate(columns, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill = HEADER_FILL
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width


def write_rows(ws, rows, wrap_from=2):
    for ri, row in enumerate(rows, start=2):
        for ci, v in enumerate(row, start=1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = BORDER
            c.alignment = Alignment(vertical='top', wrap_text=(ci >= wrap_from))
            c.font = Font(size=10)


def build_procedure():
    wb = Workbook()
    ws = wb.active
    ws.title = '手順書'
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 100
    lines = [
        ('title', 'English Speaking Drill 外注テスト手順書'),
        ('plain', f'作成日: {TODAY} ／ 作成者: 藤岡'),
        ('plain', ''),
        ('sec', '1. このテストの目的'),
        ('plain', '英会話ゲーム「English Speaking Drill」が仕様どおりに動くこと、'
                  'および直近の改修（音声・画面・ログインまわり）が直っていることを確認します。'),
        ('plain', 'テストする内容はすべて「外注テスト項目書」に書いてあります。'
                  '専門知識は不要です。書かれた手順どおりに操作して、結果を記入してください。'),
        ('plain', ''),
        ('sec', '2. 準備するもの'),
        ('plain', '・パソコン（Google Chrome 最新版）※音声認識のテストはChromeで行ってください'),
        ('plain', '・マイク（内蔵マイクでOK）とスピーカー（またはイヤホン）'),
        ('plain', '・できればiPadまたはiPhone（横画面の確認用。なければパソコンのみでOK）'),
        ('plain', '・テスト用URL・テスト用のユーザーIDとパスワード（別途お渡しします）'),
        ('plain', ''),
        ('sec', '3. はじめる前の設定'),
        ('plain', '① Chromeでテスト用URLを開く'),
        ('plain', '② 最初にマイクの許可を求められたら「許可」を押す'),
        ('plain', '③ 音が出ることを確認する（音量は中くらいに）'),
        ('plain', ''),
        ('sec', '4. テストの進め方'),
        ('plain', '① 「外注テスト項目書」を開き、上から順に1件ずつ実施します'),
        ('plain', '② 各行の「手順」どおりに操作し、「期待結果」のとおりになるか確認します'),
        ('plain', '③ 「結果」列に OK または NG を記入します'),
        ('plain', '④ NGの場合は「備考」列に、（1）何をしたか（2）何が起きたか（3）画面のスクリーンショット'
                  'の有無を書いてください。スクリーンショットはファイルで添付してもらえると助かります'),
        ('plain', '⑤ 表示が一瞬で見逃した場合などは、同じ手順をもう一度やり直してかまいません'),
        ('plain', ''),
        ('sec', '5. 注意事項'),
        ('plain', '・シート「A_機能テスト」は仕様からランダムに選んだ確認項目です'),
        ('plain', '・シート「B_要望確認テスト」は、これまでに報告のあった不具合・要望が'
                  '直っているかの確認です。全件実施してください'),
        ('plain', '・「（発生した場合のみ）」と書かれた項目は、その現象が起きたときだけ記入してください'),
        ('plain', '・音声の聞こえ方（発音など）は主観でかまいません。気になったことは何でも備考に書いてください'),
        ('plain', '・途中でゲームが完全に動かなくなった場合は、ページを再読み込み（F5）して続けてください。'
                  'その際は必ず備考に状況を書いてください'),
        ('plain', ''),
        ('sec', '6. 困ったとき'),
        ('plain', '・ログインできない／URLが開けない → 藤岡までご連絡ください'),
        ('plain', '・エラーメッセージに英数字のコード（例: AUTH-001）が出た場合は、そのコードを備考に書いてください'),
    ]
    r = 1
    for kind, text in lines:
        c = ws.cell(row=r, column=2, value=text)
        if kind == 'title':
            c.font = Font(bold=True, size=15)
        elif kind == 'sec':
            c.font = Font(bold=True, size=12, color='1E3A8A')
            c.fill = SECTION_FILL
        else:
            c.font = Font(size=10.5)
            c.alignment = Alignment(wrap_text=True, vertical='top')
        r += 1
    out = os.path.join(ROOT, 'outputs', f'外注テスト手順書_{TODAY}.xlsx')
    wb.save(out)
    return out


def build_items():
    wb = Workbook()
    wb.remove(wb.active)

    rng = random.Random(SEED)
    sampled = sorted(rng.sample(range(len(POOL)), SAMPLE_SIZE))

    ws = wb.create_sheet('A_機能テスト')
    style_header(ws, [('項目No', 8), ('対象画面', 12), ('手順', 46), ('期待結果', 46),
                      ('結果(OK/NG)', 11), ('備考', 30)])
    rows = []
    for i, idx in enumerate(sampled, start=1):
        screen, step, expect = POOL[idx]
        rows.append((f'A-{i:02d}', screen, step, expect, '', ''))
    write_rows(ws, rows, wrap_from=3)
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet('B_要望確認テスト')
    style_header(ws2, [('項目No', 8), ('元の要望', 9), ('要望の内容', 24), ('対象画面', 13),
                       ('手順', 42), ('期待結果', 42), ('結果(OK/NG)', 11), ('備考', 26)])
    rows = []
    for i, (no, summary, screen, step, expect) in enumerate(REQUESTS, start=1):
        rows.append((f'B-{i:02d}', no, summary, screen, step, expect, '', ''))
    write_rows(ws2, rows, wrap_from=3)
    ws2.auto_filter.ref = ws2.dimensions

    info = wb.create_sheet('このファイルについて', 0)
    info.column_dimensions['A'].width = 100
    notes = [
        'English Speaking Drill 外注テスト項目書',
        f'作成日: {TODAY}',
        '',
        '・進め方は「外注テスト手順書」をお読みください',
        '・A_機能テスト: 仕様書v2.0からランダムに選んだ確認項目'
        f'（{SAMPLE_SIZE}件／乱数シード{SEED}で再現可能）',
        '・B_要望確認テスト: 2026-03-14以降にご報告いただいた不具合・要望の全件確認（21件）',
        '・「結果」列に OK / NG を記入し、NGは備考に詳細をお願いします',
    ]
    for i, t in enumerate(notes, start=1):
        c = info.cell(row=i, column=1, value=t)
        c.font = Font(bold=(i == 1), size=14 if i == 1 else 10.5)

    out = os.path.join(ROOT, 'outputs', f'外注テスト項目書_{TODAY}.xlsx')
    wb.save(out)
    return out


def main():
    os.makedirs(os.path.join(ROOT, 'outputs'), exist_ok=True)
    p1 = build_procedure()
    p2 = build_items()
    print(f'生成完了:\n  {p1}\n  {p2}')
    print(f'  機能テスト: {SAMPLE_SIZE}件(プール{len(POOL)}件から抽出) / 要望確認: {len(REQUESTS)}件')


if __name__ == '__main__':
    main()
