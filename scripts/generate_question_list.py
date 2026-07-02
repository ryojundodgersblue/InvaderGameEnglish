#!/usr/bin/env python3
"""問題カタログ(人が読む用の問題一覧Excel)を生成する。

入力:
    scripts/reports/catalog_source.json … build_from_question_sheet.py が出力する
        日本語訳・シート内メモつきの問題データ
    backend/data/parts.json, answer_patterns.json

出力:
    outputs/問題カタログ_<日付>.xlsx
        学年別タブ(中1/中2/中3) + サマリータブ。オートフィルタ・デモ行ハイライトつき。
"""
import json
import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE, "backend", "data")
OUTPUT_DIR = os.path.join(BASE, "outputs")
CATALOG_SOURCE = os.path.join(BASE, "scripts", "reports", "catalog_source.json")
TODAY = os.environ.get('BUILD_DATE') or date.today().isoformat()

with open(CATALOG_SOURCE, encoding="utf-8") as f:
    catalog = json.load(f)

with open(os.path.join(DATA_DIR, "parts.json"), encoding="utf-8") as f:
    parts = {p["part_id"]: p for p in json.load(f)}

with open(os.path.join(DATA_DIR, "answer_patterns.json"), encoding="utf-8") as f:
    ans_map = {}
    for ap in json.load(f):
        ans_map.setdefault(ap["question_id"], []).append(ap["expected_text"])

# ── 行データ構築 ──────────────────────────────────────────
rows = []
for q in catalog:
    part = parts.get(q["part_id"], {})
    answers = " / ".join(ans_map.get(q["question_id"], []) or [q.get("answer", "")])
    rows.append({
        "学年": part.get("grade_id", "?"),
        "パート": part.get("part_no", "?"),
        "サブパート": part.get("subpart_no", "?"),
        "番号": q["display_order"],
        "デモ": "◎" if q["is_demo"] else "",
        "回答条件": part.get("requirement", ""),
        "問題文": q["question_text"],
        "解答(別解は「/」区切り)": answers,
        "問題文(日本語)": q.get("jp_question", ""),
        "解答(日本語)": q.get("jp_answer", ""),
        "イラスト": (q.get("image_url") or "").replace("/questions/", ""),
        "シート内メモ(DB非投入)": q.get("notes", ""),
        "question_id": q["question_id"],
    })

rows.sort(key=lambda r: (r["学年"], r["パート"], r["サブパート"], r["番号"]))

# ── Excel 生成 ──────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2563EB")
DEMO_FILL = PatternFill("solid", fgColor="DBEAFE")
ODD_FILL = PatternFill("solid", fgColor="F8FAFC")
EVEN_FILL = PatternFill("solid", fgColor="EFF6FF")
BORDER_SIDE = Side(style="thin", color="CBD5E1")
THIN_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)

COLUMNS = [
    ("パート", 7),
    ("サブパート", 9),
    ("番号", 6),
    ("デモ", 6),
    ("回答条件", 42),
    ("問題文", 38),
    ("解答(別解は「/」区切り)", 38),
    ("問題文(日本語)", 26),
    ("解答(日本語)", 26),
    ("イラスト", 17),
    ("シート内メモ(DB非投入)", 28),
    ("question_id", 11),
]
WRAP_FROM = 5  # 回答条件以降は折り返し

wb = Workbook()
wb.remove(wb.active)


def add_grade_sheet(grade_no, grade_rows):
    ws = wb.create_sheet(f"中{grade_no}")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    prev_part = None
    for row_idx, r in enumerate(grade_rows, start=2):
        is_demo = r["デモ"] == "◎"
        part_key = (r["パート"], r["サブパート"])
        fill = DEMO_FILL if is_demo else (ODD_FILL if row_idx % 2 == 0 else EVEN_FILL)
        values = [r["パート"], r["サブパート"], r["番号"], r["デモ"],
                  r["回答条件"] if part_key != prev_part else "",
                  r["問題文"], r["解答(別解は「/」区切り)"],
                  r["問題文(日本語)"], r["解答(日本語)"], r["イラスト"],
                  r["シート内メモ(DB非投入)"], r["question_id"]]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx >= WRAP_FROM))
            cell.font = Font(size=9)
        prev_part = part_key
    ws.auto_filter.ref = ws.dimensions


for g in (1, 2, 3):
    add_grade_sheet(g, [r for r in rows if r["学年"] == g])

# サマリータブ
ws2 = wb.create_sheet("サマリー", 0)
ws2.column_dimensions["A"].width = 42
ws2.column_dimensions["B"].width = 14
summary = [
    ("English Speaking Drill 問題カタログ", ""),
    (f"作成日: {TODAY}", ""),
    ("", ""),
    ("項目", "件数"),
    ("総問題数", len(rows)),
    ("デモ問題数", sum(1 for r in rows if r["デモ"])),
    ("本問題数", sum(1 for r in rows if not r["デモ"])),
    ("パート数(全)", len(parts)),
]
for g in (1, 2, 3):
    g_rows = [r for r in rows if r["学年"] == g]
    g_parts = len({(r["パート"], r["サブパート"]) for r in g_rows})
    summary.append((f"  中{g}: パート数 / 問題数", f"{g_parts} / {len(g_rows)}"))
summary += [
    ("", ""),
    ("見方", ""),
    ("・タブ「中1」「中2」「中3」が学年別の問題一覧です", ""),
    ("・「デモ」列が ◎ の行は各パート1問目のデモ問題です", ""),
    ("・「回答条件」は各パートの先頭行のみ記載しています", ""),
    ("・「シート内メモ」列は元シートの作業メモで、ゲームには反映されません", ""),
    ("・1行目のフィルタで学年・パートの絞り込みができます", ""),
]
for r_idx, (label, val) in enumerate(summary, start=1):
    c = ws2.cell(row=r_idx, column=1, value=label)
    c.font = Font(bold=(r_idx in (1, 4)), size=11 if r_idx == 1 else 10)
    ws2.cell(row=r_idx, column=2, value=val).alignment = Alignment(horizontal="right")

out_path = os.path.join(OUTPUT_DIR, f"問題カタログ_{TODAY}.xlsx")
os.makedirs(OUTPUT_DIR, exist_ok=True)
wb.save(out_path)
print(f"生成完了: {out_path}")
print(f"  総問題数: {len(rows)} / デモ: {sum(1 for r in rows if r['デモ'])}")
