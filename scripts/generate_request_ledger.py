#!/usr/bin/env python3
"""QA・要望管理ワークブック(②)を生成する。

入力:
    QUESTION_XLSX(既定: 親ディレクトリの English Speaking Drill仕様書.xlsx)
        … Q&A / 修正依頼 シート
    backend/data/*.json … 現在のDB内容(修正依頼との突合用)

出力:
    outputs/Mukaさん_QA・要望管理_<日付>.xlsx
        - QA        : 純粋な質問(仕様確認・費用・使い方など)
        - 要望       : 変更・追加の依頼。カテゴリ分類＋対応状況＋対応内容(コミット)
        - コンテンツ修正依頼 : 修正依頼シート769件＋現在のDB内容との突合列

分類はこのスクリプト内の対応表(CLASSIFY/HANDLED)で管理する。
"""
import json
import os
import re
import sys
from datetime import date

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), 'lib'))
from ids import target_to_question_id, target_to_part_id  # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
QUESTION_XLSX = os.environ.get(
    'QUESTION_XLSX',
    os.path.join(os.path.dirname(ROOT), 'English Speaking Drill仕様書.xlsx'),
)
DATA_DIR = os.path.join(ROOT, 'backend', 'data')
TODAY = os.environ.get('BUILD_DATE') or date.today().isoformat()
OUT = os.path.join(ROOT, 'outputs', f'Mukaさん_QA・要望管理_{TODAY}.xlsx')

# コミットID(短縮)。⑥トレーサビリティ用
C_IMG = '2882b73'    # イラスト116枚反映
C_DATA = '8572b78'   # DB再構築
C_TTS_BE = '98d3637'  # TTSキャッシュ版数+SSML+200ms
C_TTS_FE = 'c30d579'  # TTS頭切れ・リトライ・キャッシュ
C_FREEZE = 'fec1e55'  # フリーズ20秒+先読み
C_LAYOUT = '33562fa'  # 2カラム
C_AUTH = 'fc7d0bf'   # 認証導線+ヘルスチェック
C_MATCH = 'fb4ee3a'  # 短縮形正規化
C_PREV = '70392e8'   # 2026-03-28の既修正(TTS/スコア/タイマー/ログイン)

# --- Q&A行の分類: No → (種別, カテゴリ) ---------------------------------
# 種別: 'QA'(純粋な質問) / 'R'(要望)
# カテゴリ(要望のみ): バグ / 機能改善 / データ修正 / イラスト / その他
Q = ('QA', '')
CLASSIFY = {
    **{n: Q for n in range(1, 42)},          # 要件ヒアリング・費用・技術質問
    42: ('R', '機能改善'),                    # 爆破音を入れたい
    43: Q, 44: ('R', '機能改善'),             # 大カテゴリー選択
    45: Q, 46: ('R', '機能改善'),             # 問題数を増やしたい
    47: ('R', '機能改善'),                    # ボス単語シューティング(見送り)
    48: ('R', '機能改善'),                    # 長文出題
    49: Q, 50: ('R', '機能改善'),             # 発話ビーム(見送り)
    **{n: Q for n in (51, 52, 53, 54, 56, 57, 59, 61, 62, 64, 65, 66, 67, 68)},
    55: ('R', '機能改善'),                    # 正解の音声+テキスト表示
    58: ('R', '機能改善'),                    # クリア時の音楽など
    60: ('R', '機能改善'),                    # スコア管理者確認・途中参加
    63: ('R', '機能改善'),                    # フォント読みやすく
    **{n: Q for n in range(69, 78)},          # 問題選択/クリア条件/問題数などの取り決め
    78: ('R', '機能改善'),                    # 途中で答えたら30秒待たない
    79: ('R', '機能改善'),                    # 前学年を見れないように(見送り)
    80: Q, 81: Q,
    82: ('R', '機能改善'),                    # ニックネーム横に本名
    **{n: Q for n in (83, 84, 85, 86, 88, 89, 90, 91, 92, 93, 94, 95, 96)},
    87: ('R', '機能改善'),                    # 例題(デモ問題)を毎回
    97: ('R', '機能改善'),                    # 10回挑戦で自動クリア解除
    98: ('R', '機能改善'),                    # ゲーム名変更
    99: ('R', '機能改善'),                    # ランキングに速さ(保留)
    100: Q, 101: Q,
    102: ('R', '機能改善'),                   # 自動で次のステージ
    103: Q, 104: Q, 105: Q, 106: Q, 107: Q,
    108: ('R', 'バグ'),                       # ログイン1分
    109: ('R', '機能改善'),                   # 絵が消えるのを0.5秒短く
    110: ('R', 'イラスト'),                   # 1-5-2 イラスト差し替え
    111: ('R', 'データ修正'),                 # 1-4 Requirement
    112: ('R', 'イラスト'),                   # 1-6-1 イラスト
    113: ('R', 'バグ'),                       # あと何問でクリア誤り
    114: ('R', 'イラスト'),                   # 1-7-1 イラスト逆
    115: ('R', '機能改善'),                   # demo start 表示延長
    116: ('R', 'バグ'),                       # 問題画面から戻れない
    117: ('R', 'データ修正'),                 # 1-9-1 Requirement
    118: ('R', 'イラスト'),                   # 1-10-2 She→He
    119: ('R', 'イラスト'),                   # 1-9-2 イラスト間違い
    120: ('R', 'データ修正'),                 # 1-11-1 Requirement
    121: ('R', 'データ修正'),                 # 1-11-2 Requirement
    122: ('R', 'データ修正'),                 # 1-12 代名詞追記
    123: Q,                                   # 格納場所を教えて
    124: Q,                                   # 訂正文を書く場所
    125: ('R', 'バグ'),                       # 1-16-2-1 デモでフリーズ
    126: ('R', 'データ修正'),                 # 1-18-1-1
    127: ('R', 'データ修正'),                 # 1-20-2-1
    128: ('R', 'バグ'),                       # read過去形(1-21-1-4)
    129: ('R', 'バグ'),                       # 1-23-2-6 イラスト表示されない
    130: Q,                                   # (欠番調整用・未使用)
    # --- 2026-03-14以降(⑤の実施対象) ---
    132: ('R', 'バグ'),
    133: ('R', 'バグ'),
    134: ('R', 'バグ'),
    135: ('R', '機能改善'),
    136: ('R', '機能改善'),
    137: ('R', '機能改善'),
    138: ('R', 'バグ'),
    139: ('R', '機能改善'),
    140: ('R', '機能改善'),
    141: ('R', 'バグ'),
    142: ('R', 'バグ'),
    143: ('R', 'バグ'),
    144: ('R', 'データ修正'),
    145: ('R', 'バグ'),
    146: ('R', 'バグ'),
    147: ('R', 'バグ'),
    148: ('R', 'バグ'),
    149: ('R', '機能改善'),
}

# --- 2026-03-14以降の対応内容: No → (対応状況, 対応内容) ------------------
HANDLED = {
    132: ('対応済 ✔', f'TTS応答解析の修正で対応済({C_PREV}, 2026-03-28)。1_24_1/1_25_1の'
                     f'イラスト24枚を新規追加({C_IMG})。ローカル実機テストで回帰確認済み'),
    133: ('対応済 ✔', f'残り問題数の計算をCORRECT_TO_CLEAR定数に統一({C_PREV}, 2026-03-28)。'
                     f'回帰確認済み'),
    134: ('対応済 ✔', f'タイムアウト時のタイマー停止＋フリーズ回復ダイアログ(リトライ/次の問題へ/'
                     f'やめる)を追加({C_PREV}, 2026-03-28)。表示までの時間はNo137で20秒に短縮({C_FREEZE})'),
    135: ('対応済 ✔', f'ログインボタンに押下中の表示(Logging in…/無効化)を追加({C_PREV}, 2026-03-28)'),
    136: ('対応済 ✔', f'ログイン画面表示時にサーバーへウォームアップpingを送信し、5秒以上かかる場合'
                     f'「サーバーを起動しています…」を表示({C_AUTH})。※無料ホスティングの再起動自体を'
                     f'なくすには有料プラン等が必要(相談事項)'),
    137: ('対応済 ✔', f'フリーズ検知を30秒→20秒に短縮し定数化({C_FREEZE})'),
    138: ('対応済 ✔', f'read→現在形発音の補正は実装済みだったが、補正前に合成された音声が'
                     f'キャッシュに残り再生されていた。キャッシュに版数を導入し旧音声を無効化({C_TTS_BE})'),
    139: ('対応済 ✔', f'認証切れを検知したら自動でログイン画面へ移動し「セッションの有効期限が切れました。'
                     f'ログインし直してください。」を表示({C_AUTH})'),
    140: ('対応済 ✔', f'No139と同対応。スタート押下時の認証切れも自動でログイン画面へ誘導({C_AUTH})'),
    141: ('対応済 ✔', f'画像つき問題は問題文(左)とイラスト(右)の2カラム配置に変更し重なりを解消({C_LAYOUT})'),
    142: ('対応済 ✔', f'No138と同対応(キャッシュ版数化で旧音声を無効化)({C_TTS_BE})'),
    143: ('対応済 ✔', f'合成音声の冒頭に200msの無音を挿入({C_TTS_BE})＋再生前に読み込み完了を待つ'
                     f'({C_TTS_FE})の両面で頭切れを解消'),
    144: ('対応済 ✔', f'questionシートの3-8-2(8問)をDBに反映({C_DATA})し、イラスト8枚を追加({C_IMG})。'
                     f'※本番反映はGoogle Sheets同期時'),
    145: ('対応済 ✔', f'TTS失敗の可視化・自動リトライ・先読みで読み上げ欠落を解消({C_TTS_FE}/{C_FREEZE})。'
                     f'※「回答条件」の文面自体は音声で読まない仕様(読むのは問題文2回)'),
    146: ('対応済 ✔', f'No145と同対応({C_TTS_FE}/{C_FREEZE})'),
    147: ('対応済 ✔', f'DB再構築で全パート8問・出題順1〜8を機械検証({C_DATA})。'
                     f'※本番反映はGoogle Sheets同期時'),
    148: ('対応済 ✔', f'questionシートの3-24以降を含む全276パートをDB化({C_DATA})。'
                     f'※本番反映はGoogle Sheets同期時'),
    149: ('対応済 ✔', f'短縮形の言い換え(I\'m⇔I am等)を正解扱いにする正規化を追加({C_MATCH})。'
                     f'判定のしきい値(0.62)は変更なし'),
}
UNNUMBERED_HANDLED = [
    # (起票日, 質問内容キー, 種別, カテゴリ, 対応状況, 対応内容)
    ('認識してもらえない', 'R', '機能改善',
     '対応済 ✔', f'1語回答は認識文に単語が含まれていれば正解とする判定を追加({C_MATCH})。'
                f"解答'I'の問題に別解'eye'も追加({C_DATA})"),
    ('過去形で発音され', 'R', 'バグ',
     '対応済 ✔', f'No138と同対応。キャッシュ版数化で旧音声を無効化({C_TTS_BE})'),
    ('read を過去形で発音するレッドの文章', 'R', 'バグ',
     '対応済 ✔', f'No138と同対応({C_TTS_BE})。該当5文は本番反映後に発音を個別確認する'),
]

CUTOFF = date(2026, 3, 14)

# ---------------- スタイル ----------------
HEADER_FILL = PatternFill('solid', fgColor='2563EB')
DONE_FILL = PatternFill('solid', fgColor='DCFCE7')
TARGET_FILL = PatternFill('solid', fgColor='FEF9C3')
ODD_FILL = PatternFill('solid', fgColor='F8FAFC')
BORDER = Border(*(Side(style='thin', color='CBD5E1'),) * 4)


def style_header(ws, columns):
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 24
    for i, (name, width) in enumerate(columns, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill = HEADER_FILL
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width


def write_row(ws, row_idx, values, fill=None, wrap_from=1):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row_idx, column=i, value=v)
        if fill:
            c.fill = fill
        c.border = BORDER
        c.alignment = Alignment(vertical='top', wrap_text=(i >= wrap_from))
        c.font = Font(size=9)


def fmt_date(v):
    if v is None:
        return ''
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    return str(v)


def parse_no(v):
    if v is None:
        return None
    s = str(v).strip()
    s = s.translate(str.maketrans('０１２３４５６７８９', '0123456789'))
    m = re.match(r'^(\d+)', s)
    return int(m.group(1)) if m else None


def main():
    src = openpyxl.load_workbook(QUESTION_XLSX, data_only=True)

    # ---- Q&A rows ----
    qa_rows = []
    for r in src['Q&A'].iter_rows(min_row=4, values_only=True):
        no_raw, status, filed, asker, q_text, answerer, a_text, done = r[:8]
        if q_text is None and no_raw is None:
            continue
        qa_rows.append({
            'no': parse_no(no_raw),
            'status': str(status or '').strip(),
            'filed': filed, 'asker': str(asker or '').strip(),
            'q': str(q_text or '').strip(),
            'answerer': str(answerer or '').strip(),
            'a': str(a_text or '').strip(),
            'done': done,
        })

    wb = Workbook()
    wb.remove(wb.active)

    # ---- 表紙 ----
    cover = wb.create_sheet('はじめに')
    cover.column_dimensions['A'].width = 100
    lines = [
        'Mukaさん QA・要望管理',
        f'作成日: {TODAY}',
        '',
        '元データ: English Speaking Drill仕様書.xlsx の「Q&A」「修正依頼」シート',
        '',
        '■ シート構成',
        '・QA … 純粋な質問と回答(仕様確認・費用・使い方など)。対応は不要なもの',
        '・要望 … 変更・追加の依頼。カテゴリ(バグ/機能改善/データ修正/イラスト/その他)で分類',
        '　 - 「対応状況」が「対応済 ✔」の行は今回(2026-07)の改修で対応したもの',
        '　 - 「対応内容」の英数字7桁はGitコミットID(開発者向けの変更履歴参照用)',
        '・コンテンツ修正依頼 … 修正依頼シートの全769件',
        '　 - 問題文/解答/Requirementの修正は「questionシートが正」の方針でDBを再構築',
        '　 - 「現在のDB内容」列で反映結果を確認できます',
        '',
        '■ 今回の実施範囲(ご依頼分)',
        '・2026-03-14以降に起票されたQ&Aの要望(No.132〜149＋6/3起票分)を実施',
        '・それ以前の要望は起票時の対応状況をそのまま記載',
    ]
    for i, t in enumerate(lines, start=1):
        c = cover.cell(row=i, column=1, value=t)
        c.font = Font(bold=(i == 1), size=14 if i == 1 else 10)

    # ---- QA sheet ----
    ws_qa = wb.create_sheet('QA')
    style_header(ws_qa, [('No', 6), ('起票日', 11), ('質問者', 9), ('質問内容', 55),
                         ('回答者', 9), ('回答内容', 55), ('完了日', 11)])
    ri = 2
    for row in qa_rows:
        cls = CLASSIFY.get(row['no'], ('QA', ''))
        if row['no'] is not None and cls[0] != 'QA':
            continue
        if row['no'] is None:
            continue  # 未採番行は要望として扱う
        write_row(ws_qa, ri, [row['no'], fmt_date(row['filed']), row['asker'], row['q'],
                              row['answerer'], row['a'], fmt_date(row['done'])],
                  fill=ODD_FILL if ri % 2 == 0 else None, wrap_from=4)
        ri += 1
    ws_qa.auto_filter.ref = ws_qa.dimensions

    # ---- 要望 sheet ----
    ws_r = wb.create_sheet('要望')
    style_header(ws_r, [('No', 6), ('起票日', 11), ('カテゴリ', 11), ('要望内容', 48),
                        ('当時の回答', 34), ('対応状況', 13), ('対応内容(コミットID付き)', 52),
                        ('完了日', 11)])
    ri = 2
    unnumbered_idx = 0
    for row in qa_rows:
        no = row['no']
        if no is not None:
            cls = CLASSIFY.get(no)
            if not cls or cls[0] != 'R':
                continue
            category = cls[1]
            filed = row['filed']
            is_target = filed and hasattr(filed, 'date') and filed.date() >= CUTOFF
            if no in HANDLED:
                status, action = HANDLED[no]
            elif str(row['status']).startswith('完了'):
                status, action = '完了(起票時対応)', ''
            elif row['status']:
                status, action = row['status'], ''
            else:
                status, action = '未対応', ''
            fill = DONE_FILL if no in HANDLED else (TARGET_FILL if is_target else None)
            write_row(ws_r, ri, [no, fmt_date(filed), category, row['q'], row['a'],
                                 status, action, fmt_date(row['done'])],
                      fill=fill, wrap_from=4)
            ri += 1
        else:
            # 未採番(6/3)行
            text = row['q']
            matched = None
            for key, typ, cat, status, action in UNNUMBERED_HANDLED:
                if key in text:
                    matched = (cat, status, action)
                    break
            cat, status, action = matched if matched else ('その他', '要確認', '')
            write_row(ws_r, ri, ['(未採番)', fmt_date(row['filed']), cat, text, row['a'],
                                 status, action, ''],
                      fill=DONE_FILL if matched else TARGET_FILL, wrap_from=4)
            unnumbered_idx += 1
            ri += 1
    ws_r.auto_filter.ref = ws_r.dimensions

    # ---- コンテンツ修正依頼 sheet ----
    with open(os.path.join(DATA_DIR, 'questions.json'), encoding='utf-8') as f:
        q_by_id = {q['question_id']: q for q in json.load(f)}
    with open(os.path.join(DATA_DIR, 'answer_patterns.json'), encoding='utf-8') as f:
        a_by_qid = {}
        for a in json.load(f):
            a_by_qid.setdefault(a['question_id'], []).append(a['expected_text'])
    with open(os.path.join(DATA_DIR, 'parts.json'), encoding='utf-8') as f:
        p_by_id = {p['part_id']: p for p in json.load(f)}
    deployed_images = set(os.listdir(os.path.join(ROOT, 'frontend', 'public', 'questions')))

    ws_c = wb.create_sheet('コンテンツ修正依頼')
    style_header(ws_c, [('記入日', 11), ('対象', 13), ('修正種別', 15), ('修正内容', 46),
                        ('起票時の状況', 11), ('担当', 8), ('対応状況(今回)', 22),
                        ('現在のDB内容(問題/解答/条件)', 55)])
    ri = 2
    for r in src['修正依頼'].iter_rows(min_row=3, values_only=True):
        filed, target, kind, content, status = r[0], r[1], r[2], r[3], r[4]
        worker = r[5] if len(r) > 5 else None
        if target is None and content is None:
            continue
        kind_s = str(kind or '').strip()
        qid = target_to_question_id(target)
        pid = target_to_part_id(target)
        db_now = ''
        handled = ''
        if kind_s in ('問題文の修正', '解答のみ修正', 'アポストロフィの統一'):
            handled = f'DB再構築で反映({C_DATA})※questionシート準拠'
            if qid and qid in q_by_id:
                q = q_by_id[qid]
                ans = ' / '.join(a_by_qid.get(qid, []))
                db_now = f"問題: {q['question_text']}\n解答: {ans}"
            elif pid and pid in p_by_id:
                db_now = '(対象問題を特定できず。パートは存在)'
            else:
                db_now = '(対象を特定できず)'
        elif kind_s == 'Requirementの修正':
            handled = f'DB再構築で反映({C_DATA})※questionシート準拠'
            if pid and pid in p_by_id:
                db_now = f"条件: {p_by_id[pid]['requirement']}"
            else:
                db_now = '(対象パートを特定できず)'
        elif kind_s == 'イラスト差し替え':
            m = re.findall(r'\d+', str(target or ''))
            fn_prefix = '_'.join(m[:3]) if len(m) >= 3 else None
            if fn_prefix:
                candidates = [f for f in deployed_images if f.startswith(fn_prefix + '_')]
                if len(m) >= 4:
                    fn = f"{'_'.join(m[:3])}_{int(m[3]):02d}.png"
                    ok = fn in deployed_images
                    handled = f'イラスト反映済({C_IMG})' if ok else '画像ファイル未確認'
                    db_now = f'ファイル: {fn}' + ('' if ok else ' (存在せず)')
                else:
                    handled = f'イラスト反映済({C_IMG})' if candidates else '画像ファイル未確認'
                    db_now = f'該当画像 {len(candidates)}枚を配置済み'
            else:
                handled = '対象不明・要確認'
        else:
            handled = '個別確認'
        fill = DONE_FILL if handled.startswith(('DB再構築', 'イラスト反映')) else None
        write_row(ws_c, ri, [fmt_date(filed), str(target or ''), kind_s,
                             str(content or ''), str(status or ''), str(worker or ''),
                             handled, db_now],
                  fill=fill, wrap_from=4)
        ri += 1
    ws_c.auto_filter.ref = ws_c.dimensions

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f'生成完了: {OUT}')
    print(f"  QA: {ws_qa.max_row - 1}行 / 要望: {ws_r.max_row - 1}行 / "
          f"コンテンツ修正依頼: {ws_c.max_row - 1}行")


if __name__ == '__main__':
    main()
