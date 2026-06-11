#!/usr/bin/env python3
"""完成版データセットの構築。

入力(変更しない):
    outputs/Data Specs_2026-06-11.xlsx      ベースデータ(正本)
    outputs/Mukaさん_ユーザー要望書.xlsx     コンテンツ修正依頼(未対応673件)

処理:
    1. ベースデータの型正規化(id→int, bool統一, 日付→'YYYY/MM/DD HH:MM:SS')
    2. 修正依頼の適用(全件をapplied/skipped/already_ok/supersededで記録)
       - その他(3): image_urlタイポ修正
       - 問題文の修正(346): 以下の形式を判別して適用
         * 「問題文：Q 解答：A」形式 → Q/Aペア置換
         * 「問題文、解答ともに訂正 Q A」形式 → 2文に分割してQ/Aペア置換
         * 「Q<TAB>A」「Q → A」「Q/A」形式 → Q/Aペア置換
         * それ以外 → question_text置換
           (アポストロフィ違いのみで解答文に一致する場合は解答側へルーティング)
       - 新規サブパート追加: 存在しないpartへの問題文修正が8問完全セットで
         揃っている場合はpart+questions+answer_patternsを新設(3-31-3, 3-31-4)
       - 解答のみ修正(99): expected_text置換
       - Requirementの修正(74): parts.requirement置換
       - アポストロフィの統一(150): 全体正規化でカバーされることを照合検証
         (別の修正で文面が変わった行は superseded として記録)
       - 全体正規化: ’ ‘ → ' / ” “ → " (question_text/expected_text/requirement)
       - 3-8-2の8問にimage_url設定
    3. 出力
       - backend/data/{users,grades,parts,questions,answer_patterns,scores}.json
       - outputs/Data Specs_完成版_<日付>.xlsx
       - scripts/reports/apply_report.md / apply_report.json
"""
import datetime
import json
import os
import re
import sys
import unicodedata

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__))))
from lib import ids  # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE_XLSX = os.path.join(ROOT, 'outputs', 'Data Specs_2026-06-11.xlsx')
REQUESTS_XLSX = os.path.join(ROOT, 'outputs', 'Mukaさん_ユーザー要望書.xlsx')
DATA_DIR = os.path.join(ROOT, 'backend', 'data')
REPORT_MD = os.path.join(ROOT, 'scripts', 'reports', 'apply_report.md')
REPORT_JSON = os.path.join(ROOT, 'scripts', 'reports', 'apply_report.json')
TODAY = '2026-06-11'
FINAL_XLSX = os.path.join(ROOT, 'outputs', f'Data Specs_完成版_{TODAY}.xlsx')

CURLY = {'‘': "'", '’': "'", '“': '"', '”': '"'}

# 出力スキーマ(backend/src/utils/sheets.js の HEADERS と一致させる)
SCHEMA = {
    'users': ['id', 'user_id', 'password', 'nickname', 'real_name', 'current_grade',
              'current_part', 'current_subpart', 'is_admin', 'created_at', 'updated_at'],
    'grades': ['id', 'grade_no'],
    'parts': ['part_id', 'grade_id', 'part_no', 'subpart_no', 'requirement'],
    'questions': ['question_id', 'part_id', 'display_order', 'is_demo', 'question_text', 'image_url'],
    'answer_patterns': ['id', 'question_id', 'expected_text'],
    'scores': ['score_id', 'user_id', 'part_id', 'scores', 'clear', 'play_date', 'avg_answer_time'],
}


# ---------- 正規化ヘルパー ----------

def to_bool(v):
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ('true', '1')


def to_int(v):
    return int(str(v).strip())


def fmt_dt(v):
    """datetime/文字列 → 'YYYY/MM/DD HH:MM:SS'。パース不能な文字列はそのまま。"""
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y/%m/%d %H:%M:%S')
    s = str(v or '').strip()
    if not s:
        return ''
    m = re.fullmatch(r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})[T ]?(\d{1,2}):(\d{2}):(\d{2})(?:\.\d+)?Z?', s)
    if m:
        y, mo, d, h, mi, se = m.groups()
        return f'{y}/{int(mo):02d}/{int(d):02d} {int(h):02d}:{int(mi):02d}:{int(se):02d}'
    m = re.fullmatch(r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})', s)
    if m:
        y, mo, d = m.groups()
        return f'{y}/{int(mo):02d}/{int(d):02d} 00:00:00'
    return s


def straighten(s):
    """カーリー引用符を直線に統一する。"""
    if s is None:
        return s
    for k, v in CURLY.items():
        s = s.replace(k, v)
    return s


def text_key(s):
    """アポストロフィ・引用符・空白の揺れを無視した比較キー。"""
    if s is None:
        return ''
    s = straighten(unicodedata.normalize('NFKC', str(s)))
    return re.sub(r'\s+', ' ', s).strip()


def has_cjk(s):
    return bool(re.search(r'[぀-ヿ一-鿿]', s or ''))


def clean_en(s):
    """修正内容セルの英文を整える(前後空白除去・全角空白→半角・カーリー統一)。

    記入時に混入しがちな先頭の '/' や「問題修正」「解答修正」等のラベルも除去する。
    """
    s = str(s).replace('　', ' ')
    s = straighten(s)
    s = re.sub(r'^[/→\s]+', '', s)
    s = re.sub(r'^(問題文?(の)?修正|解答(のみ)?(の)?修正)[：:\s]*', '', s)
    return re.sub(r'\s+', ' ', s).strip()


def split_pair(raw):
    """修正内容から (新問題文, 新解答) ペアを抽出する。ペアでなければ None。

    対応形式(優先順):
      1. 「問題文：Q 解答：A」
      2. 「Q?(改行)A: 解答」
      3. 「問題文、解答ともに訂正 Q A」(2文に分割)
      4. タブ区切り「Q\\tA」
      5. 矢印区切り「Q → A」(言い換え問題: 左=出題文, 右=解答文)
      6. スラッシュ区切り「Q/A」
    """
    text = str(raw)
    m = re.search(r'問題文[：:]\s*(.+?)\s*解答[：:]\s*(.+?)(?:\s*へ変更.*)?$',
                  text.replace('\n', ' '))
    if m:
        return clean_en(m.group(1)), clean_en(m.group(2))
    m = re.match(r'^(.+?[?.!])\s*\n\s*A[：:]\s*(.+)$', text, re.S)
    if m:
        return clean_en(m.group(1)), clean_en(m.group(2))
    m = re.match(r'^\s*問題文?、?\s*解答ともに訂正[：:\s　]*(.+)$', text.replace('\n', ' '))
    if m:
        body = m.group(1).strip()
        m2 = re.match(r'^(.+?[.?!])\s+(.+)$', body)
        if m2:
            return clean_en(m2.group(1)), clean_en(m2.group(2))
    for sep in ('\t', '→', '/'):
        if sep in text:
            left, right = text.split(sep, 1)
            left, right = clean_en(left), clean_en(right)
            if left and right:
                return left, right
            return None
    return None


# ---------- ベースデータ読み込み ----------

def load_base():
    wb = openpyxl.load_workbook(BASE_XLSX, read_only=True)

    def rows_of(sheet):
        out = []
        for r in list(wb[sheet].iter_rows(values_only=True))[1:]:
            if r and r[0] is not None:
                out.append(r)
        return out

    users = []
    for r in rows_of('users'):
        users.append({
            'id': to_int(r[0]), 'user_id': str(r[1]).strip(), 'password': str(r[2]).strip(),
            'nickname': str(r[3] or '').strip(), 'real_name': str(r[4] or '').strip(),
            'current_grade': to_int(r[5]), 'current_part': to_int(r[6]), 'current_subpart': to_int(r[7]),
            'is_admin': to_bool(r[8]), 'created_at': fmt_dt(r[9]), 'updated_at': fmt_dt(r[10]),
        })
    grades = [{'id': to_int(r[0]), 'grade_no': to_int(r[1])} for r in rows_of('grades')]
    parts = []
    for r in rows_of('parts'):
        parts.append({
            'part_id': to_int(r[0]), 'grade_id': to_int(r[1]), 'part_no': to_int(r[2]),
            'subpart_no': to_int(r[3]), 'requirement': str(r[4] or '').strip(),
        })
    questions = []
    for r in rows_of('questions'):
        questions.append({
            'question_id': to_int(r[0]), 'part_id': to_int(r[1]), 'display_order': to_int(r[2]),
            'is_demo': to_bool(r[3]), 'question_text': str(r[4] or '').strip(),
            'image_url': str(r[5] or '').strip(),
        })
    answers = []
    for r in rows_of('answer_patterns'):
        answers.append({
            'id': to_int(r[0]), 'question_id': to_int(r[1]),
            'expected_text': str(r[2] or '').strip(),
        })
    scores = []
    for r in rows_of('scores'):
        scores.append({
            'score_id': to_int(r[0]), 'user_id': str(r[1]).strip(), 'part_id': to_int(r[2]),
            'scores': to_int(r[3] or 0), 'clear': to_bool(r[4]), 'play_date': fmt_dt(r[5]),
            'avg_answer_time': None,
        })
    wb.close()
    return {'users': users, 'grades': grades, 'parts': parts,
            'questions': questions, 'answer_patterns': answers, 'scores': scores}


# ---------- 修正依頼読み込み ----------

def load_requests():
    wb = openpyxl.load_workbook(REQUESTS_XLSX, read_only=True)
    ws = wb['コンテンツ修正依頼']
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for i, r in enumerate(rows[3:], start=4):  # 1始まり行番号(ヘッダ3行)
        if not r or (r[0] is None and r[1] is None):
            continue
        status = str(r[4] or '').strip()
        if status == '完了':
            continue
        out.append({
            'xlsx_row': i + 1 if rows[0] and rows[0][0] else i,  # openpyxlの実行行
            'row': i,
            'date': fmt_dt(r[0])[:10],
            'target': str(r[1]).strip() if r[1] is not None else '',
            'type': str(r[2] or '').strip(),
            'content': r[3] if r[3] is not None else '',
        })
    wb.close()
    return out


# ---------- 適用エンジン ----------

class Applier:
    def __init__(self, data):
        self.data = data
        self.q_by_id = {q['question_id']: q for q in data['questions']}
        self.a_by_qid = {}
        for a in data['answer_patterns']:
            self.a_by_qid.setdefault(a['question_id'], []).append(a)
        self.p_by_id = {p['part_id']: p for p in data['parts']}
        self.entries = []
        self.touched_qids = {}  # question_id -> [types] (overlap検出)
        self.created_parts = []
        # アポストロフィ統一行の照合用に、修正適用前の原文を保持する
        self.original_q = {q['question_id']: q['question_text'] for q in data['questions']}
        self.original_a = {a['id']: a['expected_text'] for a in data['answer_patterns']}

    def create_missing_parts(self, reqs):
        """存在しないpartを対象とする問題文修正が8問完全セットで揃っている場合、
        part / questions / answer_patterns の空エントリを新設する(新規サブパート追加)。
        """
        missing = {}  # part_id -> set(display_order)
        for req in reqs:
            if req['type'] != '問題文の修正':
                continue
            comps = ids.parse_components(ids.normalize_target(req['target']))
            if len(comps) != 4:
                continue
            pid = ids.part_id(comps[0], comps[1], comps[2])
            if pid not in self.p_by_id:
                missing.setdefault(pid, set()).add(comps[3])

        next_aid = max(a['id'] for a in self.data['answer_patterns']) + 1
        for pid in sorted(missing):
            orders = missing[pid]
            if orders != set(range(1, 9)):
                # 不完全セットは作成しない(各行は対象不明としてskipされる)
                continue
            grade = pid // 1000
            part_no = (pid % 1000) // 10
            subpart_no = pid % 10
            part = {'part_id': pid, 'grade_id': grade, 'part_no': part_no,
                    'subpart_no': subpart_no, 'requirement': ''}
            self.data['parts'].append(part)
            self.p_by_id[pid] = part
            for order in range(1, 9):
                qid = pid * 10 + order
                q = {'question_id': qid, 'part_id': pid, 'display_order': order,
                     'is_demo': order == 1, 'question_text': '', 'image_url': ''}
                self.data['questions'].append(q)
                self.q_by_id[qid] = q
                a = {'id': next_aid, 'question_id': qid, 'expected_text': ''}
                next_aid += 1
                self.data['answer_patterns'].append(a)
                self.a_by_qid[qid] = [a]
            self.created_parts.append(f'{grade}-{part_no}-{subpart_no} (part_id={pid})')

    def log(self, req, action, field='', before='', after='', reason=''):
        self.entries.append({
            'row': req['row'], 'date': req['date'], 'target': req['target'],
            'type': req['type'], 'action': action, 'field': field,
            'before': before, 'after': after, 'reason': reason,
        })

    def touch(self, qid, req):
        self.touched_qids.setdefault(qid, []).append(f"{req['type']}(row{req['row']})")

    # --- 各修正種別 ---

    def apply_image_typos(self, req):
        """その他: イラスト表示不具合 = image_urlのタイポ。固定マッピング。"""
        fixes = {
            '1-23-2-6': (12326, '/questions/1_23_2_06.png'),
            '1-28-1-5': (12815, '/questions/1_28_1_05.png'),
            '1 _41_2_01': (14121, '/questions/1_41_2_01.png'),
        }
        key = req['target']
        if key not in fixes:
            self.log(req, 'skipped', reason='その他: 対応マッピングなし(要確認)')
            return
        qid, url = fixes[key]
        q = self.q_by_id[qid]
        if q['image_url'] == url:
            self.log(req, 'already_ok', 'image_url', q['image_url'], url,
                     'ベースデータで修正済み')
        else:
            self.log(req, 'applied', 'image_url', q['image_url'], url, 'タイポ修正')
            q['image_url'] = url
        self.touch(qid, req)

    def apply_question_fix(self, req):
        qid = ids.target_to_question_id(req['target'])
        if qid is None or qid not in self.q_by_id:
            self.log(req, 'skipped', reason=f'対象を特定できない: {req["target"]!r}')
            return
        q = self.q_by_id[qid]
        answers = self.a_by_qid.get(qid, [])

        # ペア形式(問題文：/タブ/矢印/スラッシュ等) → 問題文+解答を置換
        pair = split_pair(req['content'])
        if pair:
            self._set_question_text(req, q, pair[0])
            self._set_answer_text(req, qid, answers, pair[1])
            return

        # 単独置換: 解答文側にのみ(アポストロフィ違い等で)一致する場合は解答へ
        content = clean_en(req['content'])
        if (answers and text_key(content) != text_key(q['question_text'])
                and any(text_key(content) == text_key(a['expected_text']) for a in answers)):
            self._set_answer_text(req, qid, answers, content,
                                  reason='内容が解答文に一致するため解答側へ適用')
            return
        self._set_question_text(req, q, content)

    def apply_answer_fix(self, req):
        qid = ids.target_to_question_id(req['target'])
        if qid is None or qid not in self.q_by_id:
            self.log(req, 'skipped', reason=f'対象を特定できない: {req["target"]!r}')
            return
        content = re.sub(r'^解答の修正[：:\s　]*', '', str(req['content']))
        self._set_answer_text(req, qid, self.a_by_qid.get(qid, []), clean_en(content))

    def apply_requirement_fix(self, req):
        pid = ids.target_to_part_id(req['target'])
        if pid is None or pid not in self.p_by_id:
            self.log(req, 'skipped',
                     reason=f'対象partを特定できない(要確認): {req["target"]!r}')
            return
        p = self.p_by_id[pid]
        new = straighten(str(req['content']).strip())
        if p['requirement'] == new:
            self.log(req, 'already_ok', 'requirement', p['requirement'], new)
        else:
            self.log(req, 'applied', 'requirement', p['requirement'], new)
            p['requirement'] = new

    def apply_apostrophe(self, req):
        """アポストロフィ統一: 全体正規化でカバーされるかを照合検証する。

        修正内容は「直線アポストロフィに直した同一文」。問題文/解答文のどちらに
        該当するかを照合し、一致すればその場で直線化(=全体正規化と同結果)。
        どちらにも一致しない場合は実質的な文変更の可能性があるため要確認。
        """
        qid = ids.target_to_question_id(req['target'])
        if qid is None or qid not in self.q_by_id:
            self.log(req, 'skipped', reason=f'対象を特定できない: {req["target"]!r}')
            return
        q = self.q_by_id[qid]
        answers = self.a_by_qid.get(qid, [])
        content = clean_en(req['content'])
        key = text_key(content)
        matched = []
        if key == text_key(q['question_text']):
            before = q['question_text']
            q['question_text'] = content
            matched.append(('question_text', before))
        for a in answers:
            if key == text_key(a['expected_text']):
                before = a['expected_text']
                a['expected_text'] = content
                matched.append(('expected_text', before))
        if matched:
            for field, before in matched:
                action = 'applied' if before != content else 'already_ok'
                self.log(req, action, field, before, content, 'アポストロフィ統一')
            self.touch(qid, req)
            return
        # 不一致の場合: 修正適用前の原文に一致するなら、先行する内容修正に
        # 置き換えられたケース(アポストロフィ行は原文ベースで起票されたもの)
        orig_keys = [text_key(self.original_q.get(qid, ''))]
        orig_keys += [text_key(self.original_a.get(a['id'], '')) for a in answers]
        if key in orig_keys:
            others = [t for t in self.touched_qids.get(qid, [])]
            self.log(req, 'superseded',
                     reason=f'同一問題への内容修正({", ".join(others) or "別修正"})が優先。'
                            'アポストロフィは全体正規化で統一済み')
            return
        self.log(req, 'skipped', reason='問題文・解答文のいずれにも一致せず(要確認)')

    def apply_illustration(self, req):
        self.log(req, 'skipped',
                 reason='差し替えイラストがローカル未着(クライアント要確認)')

    # --- 共通setter ---

    def _set_question_text(self, req, q, new):
        if q['question_text'] == new:
            self.log(req, 'already_ok', 'question_text', q['question_text'], new)
        else:
            self.log(req, 'applied', 'question_text', q['question_text'], new,
                     'CJK混入注意' if has_cjk(new) else '')
            q['question_text'] = new
        self.touch(q['question_id'], req)

    def _set_answer_text(self, req, qid, answers, new, reason=''):
        if not answers:
            self.log(req, 'skipped', 'expected_text', reason='回答パターンが存在しない')
            return
        a = answers[0]  # 全問題が1回答パターン(検証済み)
        if a['expected_text'] == new:
            self.log(req, 'already_ok', 'expected_text', a['expected_text'], new, reason)
        else:
            extra = 'CJK混入注意' if has_cjk(new) else ''
            self.log(req, 'applied', 'expected_text', a['expected_text'], new,
                     ('; '.join(x for x in [reason, extra] if x)))
            a['expected_text'] = new
        self.touch(qid, req)

    # --- 仕上げ ---

    def global_normalize(self):
        count = 0
        for q in self.data['questions']:
            s = straighten(q['question_text'])
            if s != q['question_text']:
                q['question_text'] = s
                count += 1
        for a in self.data['answer_patterns']:
            s = straighten(a['expected_text'])
            if s != a['expected_text']:
                a['expected_text'] = s
                count += 1
        for p in self.data['parts']:
            s = straighten(p['requirement'])
            if s != p['requirement']:
                p['requirement'] = s
                count += 1
        return count

    def fix_data_anomalies(self):
        """ベースデータ固有の不整合を補正する(適用ログに記録)。"""
        fixed = []
        # 1-16-2 (part 1162): 全278パート中唯一デモ問題が無い。
        # 他の全パートと同じく1問目をデモにする(No127でデモの存在が前提とされている)。
        q = self.q_by_id.get(11621)
        if q and not q['is_demo']:
            q['is_demo'] = True
            fixed.append('1-16-2-1 (question_id=11621): is_demo false→true (全パート共通の1問目デモ構成に統一)')
        return fixed

    def set_382_images(self):
        applied = 0
        for order in range(1, 9):
            qid = 30820 + order
            q = self.q_by_id.get(qid)
            if q is None:
                continue
            url = f'/questions/3_8_2_{order:02d}.png'
            if q['image_url'] != url:
                q['image_url'] = url
                applied += 1
        return applied


# ---------- 出力 ----------

def write_json(data):
    os.makedirs(DATA_DIR, exist_ok=True)
    for name, rows in data.items():
        path = os.path.join(DATA_DIR, f'{name}.json')
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(rows, f, ensure_ascii=False, indent=2)
            f.write('\n')


def write_xlsx(data):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name in ['users', 'grades', 'parts', 'questions', 'answer_patterns', 'scores']:
        ws = wb.create_sheet(name)
        cols = SCHEMA[name]
        ws.append(cols)
        for row in data[name]:
            ws.append([row[c] for c in cols])
    wb.save(FINAL_XLSX)


def write_report(applier, summary):
    entries = applier.entries
    os.makedirs(os.path.dirname(REPORT_MD), exist_ok=True)
    with open(REPORT_JSON, 'w', encoding='utf-8') as f:
        json.dump({'summary': summary, 'entries': entries}, f, ensure_ascii=False, indent=2)
        f.write('\n')

    by_action = {}
    for e in entries:
        by_action.setdefault(e['action'], []).append(e)
    overlaps = {qid: types for qid, types in applier.touched_qids.items() if len(types) > 1}

    def esc(s):
        return str(s).replace('|', '\\|').replace('\n', ' ⏎ ')

    with open(REPORT_MD, 'w', encoding='utf-8') as f:
        f.write(f'# 修正依頼 適用レポート ({TODAY})\n\n')
        f.write('## サマリ\n\n')
        for k, v in summary.items():
            f.write(f'- {k}: {v}\n')
        f.write('\n## クライアント(Mukaさん)要確認\n\n')
        f.write('| 項目 | 内容 |\n|---|---|\n')
        f.write('| 2-35-1-2 イラスト差し替え | 「New Year\'s Day」の差し替えイラストがローカル未着。受領後に反映要 |\n')
        f.write('| Requirement修正の対象不明1件 | 「to不定詞を使って〜」の対象が3-3-1/3-3-2/3-3-3のいずれか特定不能(記入時にExcelが日付化け) |\n')
        f.write('| 1-24-2 のイラスト | 不具合報告No132で言及があるが、イラスト未提供のためデータ上image_urlなし(画像なしで動作はする) |\n')
        f.write('| 1-44-1 の問題文 | 問題文なし(イラストのみ)。要望No102で「問題文なしでも動く」と確認済みのため現状維持 |\n')
        if applier.created_parts:
            f.write('\n## 新規サブパート作成(修正依頼の8問完全セットから)\n\n')
            for p in applier.created_parts:
                f.write(f'- {p}: Requirement+8問+解答を新設(イラストなし)\n')
        f.write('\n## スキップ(要確認)一覧\n\n')
        f.write('| 行 | 対象 | 種別 | 理由 |\n|---|---|---|---|\n')
        for e in by_action.get('skipped', []):
            f.write(f"| {e['row']} | {esc(e['target'])} | {e['type']} | {esc(e['reason'])} |\n")
        if by_action.get('superseded'):
            f.write('\n## superseded(より新しい内容修正が優先された行)\n\n')
            f.write('| 行 | 対象 | 種別 | 理由 |\n|---|---|---|---|\n')
            for e in by_action.get('superseded', []):
                f.write(f"| {e['row']} | {esc(e['target'])} | {e['type']} | {esc(e['reason'])} |\n")
        if overlaps:
            f.write('\n## 同一問題への複数修正(後勝ちで適用)\n\n')
            for qid, types in sorted(overlaps.items()):
                f.write(f'- question_id={qid}: {", ".join(types)}\n')
        f.write('\n## 全適用ログ\n\n')
        f.write('| 行 | 対象 | 種別 | 結果 | フィールド | 変更前 | 変更後 | 備考 |\n')
        f.write('|---|---|---|---|---|---|---|---|\n')
        for e in entries:
            f.write(f"| {e['row']} | {esc(e['target'])} | {e['type']} | {e['action']} | "
                    f"{e['field']} | {esc(e['before'])} | {esc(e['after'])} | {esc(e['reason'])} |\n")


def main():
    data = load_base()
    reqs = load_requests()
    applier = Applier(data)
    applier.create_missing_parts(reqs)

    order = {'その他': 0, '問題文の修正': 1, '解答のみ修正': 2,
             'Requirementの修正': 3, 'アポストロフィの統一': 4, 'イラスト差し替え': 5}
    handler = {
        'その他': applier.apply_image_typos,
        '問題文の修正': applier.apply_question_fix,
        '解答のみ修正': applier.apply_answer_fix,
        'Requirementの修正': applier.apply_requirement_fix,
        'アポストロフィの統一': applier.apply_apostrophe,
        'イラスト差し替え': applier.apply_illustration,
    }
    for req in sorted(reqs, key=lambda r: (order.get(r['type'], 9), r['row'])):
        h = handler.get(req['type'])
        if h is None:
            applier.log(req, 'skipped', reason=f'未知の修正種別: {req["type"]!r}')
            continue
        h(req)

    normalized = applier.global_normalize()
    img382 = applier.set_382_images()
    anomalies_fixed = applier.fix_data_anomalies()

    # 新規サブパート追加後もシート上の並びをID順に保つ
    data['parts'].sort(key=lambda p: p['part_id'])
    data['questions'].sort(key=lambda q: q['question_id'])
    data['answer_patterns'].sort(key=lambda a: a['id'])

    from collections import Counter
    actions = Counter(e['action'] for e in applier.entries)
    types = Counter(f"{e['type']}:{e['action']}" for e in applier.entries)
    summary = {
        '修正依頼(未対応)件数': len(reqs),
        '適用ログ件数': len(applier.entries),
        'applied': actions.get('applied', 0),
        'already_ok': actions.get('already_ok', 0),
        'superseded(別修正が優先)': actions.get('superseded', 0),
        'skipped(要確認)': actions.get('skipped', 0),
        '新規サブパート作成': applier.created_parts,
        '全体アポストロフィ正規化での追加修正フィールド数': normalized,
        '3-8-2 image_url設定': img382,
        'ベースデータ補正': anomalies_fixed,
        'データ件数': {k: len(v) for k, v in data.items()},
        '種別内訳': dict(sorted(types.items())),
    }

    write_json(data)
    write_xlsx(data)
    write_report(applier, summary)

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    skipped = [e for e in applier.entries if e['action'] == 'skipped']
    print(f"\nスキップ {len(skipped)}件:")
    for e in skipped:
        print(f"  row{e['row']} {e['target']!r} [{e['type']}] {e['reason']}")
    cjk = [e for e in applier.entries if 'CJK' in (e['reason'] or '')]
    if cjk:
        print(f"\nCJK混入注意 {len(cjk)}件:")
        for e in cjk:
            print(f"  row{e['row']} {e['target']!r} -> {e['after']!r}")
    return 0


if __name__ == '__main__':
    sys.exit(main())
