#!/usr/bin/env python3
"""English Speaking Drill 仕様書 v2.0 を生成する(①)。

現行ソース(fix/user-requests-202607ブランチの最終実装)を正として、
v1.2と同じシート構成(Cover→Revision→Index→Function List→Function Specs→
UI Specs→Data Specs→Tech Stack)で全面改訂する。

出力: outputs/English Speaking Drill仕様書_v2.0.xlsx
"""
import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TODAY = os.environ.get('BUILD_DATE') or date.today().isoformat()
OUT = os.path.join(ROOT, 'outputs', 'English Speaking Drill仕様書_v2.0.xlsx')

TITLE = 'ゲームで英語'
HEADER_FILL = PatternFill('solid', fgColor='2563EB')
TITLE_FONT = Font(bold=True, size=14, color='1E3A8A')
SUB_FONT = Font(bold=True, size=11, color='334155')
BORDER = Border(*(Side(style='thin', color='CBD5E1'),) * 4)


def new_sheet(wb, name, subtitle, columns=None):
    ws = wb.create_sheet(name)
    ws.cell(row=1, column=1, value=TITLE).font = TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = SUB_FONT
    if columns:
        for i, (label, width) in enumerate(columns, start=1):
            c = ws.cell(row=3, column=i, value=label)
            c.fill = HEADER_FILL
            c.font = Font(bold=True, color='FFFFFF', size=10)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = BORDER
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = 'A4'
    return ws


def fill_rows(ws, rows, start=4, wrap_from=1):
    for ri, row in enumerate(rows, start=start):
        for ci, v in enumerate(row, start=1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = BORDER
            c.alignment = Alignment(vertical='top', wrap_text=(ci >= wrap_from))
            c.font = Font(size=9)


def main():
    wb = Workbook()
    wb.remove(wb.active)

    # ---------------- Cover ----------------
    ws = new_sheet(wb, 'Cover', '表紙（Cover）')
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 70
    ws.cell(row=4, column=2, value='英会話ゲーム仕様書').font = Font(bold=True, size=16)
    cover_rows = [
        ('プロジェクト名', 'English Speaking Drill'),
        ('バージョン', 'v2.0'),
        ('作成日', TODAY),
        ('作成者', '藤岡諒也'),
        ('クライアント名', 'ポップ英語教室　向井真由美様'),
        ('使用技術', 'React 19 / TypeScript / Express 5 (Node.js) / Google Sheets / '
                   'Redis / Google Cloud TTS / Web Speech API'),
        ('リリース対象環境', 'Google Chrome（PC）／iPad・iPhone Safari（横画面対応）'),
        ('本番URL', 'https://invader-game-english-git-main-ryoya-fujiokas-projects.vercel.app/'),
        ('備考', '本書はv1.2以前の「予定仕様」を廃し、2026-07時点の実装を正として全面改訂した。'
               '問題データの正は questionシート→「DB_2026-07-02.xlsx」。'
               '要望の管理は「Mukaさん_QA・要望管理」ワークブックを参照。'),
    ]
    for i, (k, v) in enumerate(cover_rows, start=6):
        ws.cell(row=i, column=2, value=k).font = Font(bold=True, size=10)
        c = ws.cell(row=i, column=3, value=v)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        c.font = Font(size=10)

    # ---------------- Revision ----------------
    ws = new_sheet(wb, 'Revision', '改定履歴（Revision）',
                   [('バージョン', 12), ('日付', 12), ('修正内容', 80), ('修正者', 12)])
    fill_rows(ws, [
        ('v1.0', '2025-04-20', '初版作成', '藤岡諒也'),
        ('v1.1', '2025-05-17', 'ランキング画面、管理画面の追加', '藤岡諒也'),
        ('v1.2', '2025-07-27', 'questionシートの追加', '藤岡諒也'),
        ('v1.3', '2025-08-31', 'sampleシートの追加', '藤岡諒也'),
        ('v1.4', '2025-08-31', 'DBの再構築', '藤岡諒也'),
        ('v2.0', TODAY,
         '実装準拠で全面改訂（ソースを正とする）。2026-03-14以降起票のQ&A要望'
         '（No.132〜149、6/3起票分）を反映: TTSキャッシュ版数管理・冒頭200ms無音・'
         '読み上げリトライ/先読み、フリーズ検知20秒化、画像つき問題の2カラム化、'
         '認証切れの再ログイン導線、起動時ヘルスチェック、短縮形対応の回答正規化、'
         '3-8-2追加を含むDB再構築(276パート/2,208問)、イラスト116枚反映。', '藤岡諒也'),
    ], wrap_from=3)

    # ---------------- Index ----------------
    ws = new_sheet(wb, 'Index', '目次（Index）',
                   [('セクション名', 20), ('シート名', 18), ('説明', 70)])
    fill_rows(ws, [
        ('表紙', 'Cover', 'プロジェクト概要'),
        ('改定履歴', 'Revision', '修正履歴・バージョン管理'),
        ('機能一覧', 'Function List', '全体の機能一覧'),
        ('機能仕様詳細', 'Function Specs', '各機能の詳細設計'),
        ('画面仕様', 'UI Specs', 'UI画面の構成やUI要素'),
        ('データ仕様', 'Data Specs', 'DB（Googleスプレッドシート）の設計'),
        ('使用技術まとめ', 'Tech Stack', '使用ライブラリ・構成技術'),
        ('（別ファイル）', 'Mukaさん_QA・要望管理', 'クライアントからの質問・要望の管理台帳'),
        ('（別ファイル）', '問題カタログ', '全2,208問の一覧（学年別・日本語訳つき）'),
        ('（別ファイル）', 'DB_2026-07-02.xlsx', 'Google Sheetsへ反映するDB完成版'),
        ('（別ファイル）', '外注テスト手順書・項目書', '第三者テスト用の資材'),
    ], wrap_from=3)

    # ---------------- Function List ----------------
    ws = new_sheet(wb, 'Function List', ' 機能一覧（Function List）',
                   [('機能ID', 7), ('対象画面ID', 10), ('機能名', 18), ('処理概要', 42),
                    ('入力', 24), ('出力', 24), ('条件・制約', 32), ('エラー処理', 32), ('備考', 26)])
    fill_rows(ws, [
        ('1.0', 'G001', 'ユーザー認証',
         '・ユーザーID／パスワードの照合\n・成功時にJWTを発行しhttpOnly Cookie(authToken)に保存',
         '・ユーザーID\n・パスワード',
         '・認証結果(OK/NG)\n・ユーザー情報(進捗含む)',
         '・ユーザーIDはDB(usersシート)に登録済みであること\n・パスワードはbcryptハッシュ(平文は後方互換で比較)',
         '・不一致→「認証に失敗しました」\n・通信断→「サーバーに接続できません(NET-001)」',
         '・トークン有効期限24時間\n・ログイン試行回数の制限は未実装'),
        ('2.0', 'G001', 'サーバーウォームアップ',
         '・ログイン画面表示時に /health へ事前接続し、サーバーのコールドスタートを先行起動',
         'なし', '・疎通結果(表示なし)',
         '・失敗してもログイン処理で再接続されるため無害',
         '・ログインに5秒以上かかる場合「サーバーを起動しています…」を表示',
         '要望No136対応'),
        ('3.0', 'G002', 'ステージ選択',
         '・ユーザーの進捗(現在学年/パート/サブパート)以下のステージ一覧を返す\n・選択組み合わせの妥当性検証',
         '・ユーザーID\n・学年/パート/サブパート',
         '・選択可能なステージ一覧\n・検証結果(valid)',
         '・進捗より先のステージは選択不可\n・組み合わせがpartsシートに存在すること',
         '・存在しない組み合わせ→「無効な組み合わせです」\n・認証切れ→ログイン画面へ誘導',
         '・ゲーム開始時に無音再生で自動再生制限を解除'),
        ('4.0', 'G003', '問題出題',
         '・part_idから問題一覧を取得(出題順で最大8問)\n・1問目はデモ問題(自動で正解を提示)\n'
         '・問題文をTTSで2回読み上げ、その後回答受付',
         '・part_id',
         '・問題一覧(問題文/正解/画像URL)',
         '・1パート=8問(デモ1+本問7)\n・回答条件(Requirement)は開始前に画面表示(音声では読まない)',
         '・問題取得失敗→エラーメッセージ＋戻る導線\n・データなし→「No Data」表示',
         '・intermission中に次問題の音声・画像を先読み'),
        ('5.0', 'G003', '音声読み上げ(TTS)',
         '・問題文・正解をGoogle Cloud TTSで合成し再生\n・SSML化(XMLエスケープ+発音補正+冒頭200ms無音)\n'
         '・失敗時は自動で1回リトライ',
         '・読み上げテキスト',
         '・音声(MP3)再生',
         '・音声はRedisに24時間キャッシュ(版数つき)\n・セッション内キャッシュ最大50件\n'
         '・合成10秒/読込待ち1.5秒/再生15秒の上限',
         '・合成失敗→リトライ→失敗時はスキップしてゲーム続行(コンソールに記録)\n・認証切れ→ログイン画面へ誘導',
         '・"read"は現在形/riːd/で発音(要望No138/142)\n・頭切れ対策(要望No143)'),
        ('6.0', 'G003', '音声認識',
         '・Web Speech APIでマイク入力をリアルタイム認識\n・銃ボタンでマイクON/OFF',
         '・マイク入力',
         '・認識テキスト(複数候補を蓄積)',
         '・マイク権限が許可されていること\n・認識中は途中結果も取得(interimResults)',
         '・未対応ブラウザ→「このブラウザは音声認識に未対応です(Chrome推奨)」\n・認識停止時は自動再開(listening中)',
         '・認識結果は画面右上に「Heard: …」表示'),
        ('7.0', 'G003', '正誤判定',
         '・認識/正解テキストを正規化(小文字化・記号除去・短縮形展開)して比較\n'
         '①完全一致 ②1語正解のトークン一致 ③類似度(レーベンシュタイン/Jaccard)≧0.62',
         '・認識テキスト群\n・正解パターン群',
         '・判定結果(true/false)',
         '・短縮形(I\'m⇔I am等)はどちらの言い方でも正解\n・空文字は不正解',
         '・不正解→ミス演出後に再度回答受付(時間内)',
         '要望No149対応。閾値0.62は変更しないこと'),
        ('8.0', 'G003', 'ゲーム演出・タイマー',
         '・正解: ビーム→敵爆発→正解表示\n・不正解: 敵攻撃演出\n・制限時間30秒のカウントダウン',
         '・判定結果',
         '・アニメーション/効果音',
         '・デモ問題はタイマーなし\n・正解/時間切れ後に正解の音声+テキストを提示',
         '・時間切れ→タイマー停止→正解提示→次の問題へ',
         '・効果音: attack.mp3/miss.mp3(音量0.2)'),
        ('9.0', 'G003', 'フリーズ検知・回復',
         '・20秒間無応答を検知したら回復ダイアログを表示\n・「リトライ」「次の問題へ」「やめる」を選択可能',
         '・画面アクティビティ',
         '・回復ダイアログ',
         '・監視間隔5秒\n・finished/idle中は監視しない',
         '・「リトライ」→同一問題を再開\n・「次の問題へ」→スキップ\n・「やめる」→ステージ選択へ',
         '要望No134/137対応(30秒→20秒)'),
        ('10.0', 'G003', '中断(やめる)',
         '・ゲーム中いつでも「やめる」ボタンでステージ選択画面へ戻る',
         '・やめるボタン押下', '・画面遷移',
         '・音声・認識・タイマーを全停止してから遷移', '—', '要望No118対応'),
        ('11.0', 'G004', 'スコア計算・保存',
         '・正解数(デモ除く)をスコアとしてscoresシートに保存\n・5問以上正解でクリア',
         '・正解数/クリア判定',
         '・保存結果(score_id)',
         '・1プレイ=1レコード\n・スコアは0〜1000の範囲検証\n・ユーザーIDはJWTと一致必須',
         '・保存失敗→エラー表示(結果画面へは遷移)\n・認証切れ→ログイン画面へ誘導',
         'CORRECT_TO_CLEAR=5'),
        ('12.0', 'G004', 'ステージ進行',
         '・クリア時、または同一パートの挑戦回数が10回に到達で次のパートへ自動進行\n・usersシートの進捗を更新',
         '・ユーザーID/現在地/クリア有無',
         '・進行結果(次のパート)',
         '・進捗とDBの現在地が一致すること\n・最終パートではそれ以上進まない',
         '・進捗不一致→progress mismatch(進行せず)\n・回数不足→remaining返却',
         'REQUIRED_ATTEMPTS=10(要望No102/104対応)'),
        ('13.0', 'G005', 'ランキング',
         '・当月のスコア履歴から「挑戦回数」「正答率(平均スコア)」の上位3名を表示',
         '・なし(当月自動判定)',
         '・2種のランキング(各上位3名)',
         '・月替わりで自動リセット\n・60秒キャッシュ',
         '・データなし→空のランキング表示\n・取得失敗→「ランキング取得に失敗しました」',
         'RANKING_TOP_N=3'),
        ('14.0', 'A001', 'ユーザー管理(管理者)',
         '・新規ユーザー登録(ID自動採番・パスワード自動生成)\n・進捗(学年/パート/サブパート)の編集\n'
         '・パスワードリセット',
         '・ニックネーム/本名 ほか',
         '・ユーザーID(5桁)/初期パスワード(8桁)',
         '・管理者権限(is_admin)が必要\n・パスワードはbcryptで保存',
         '・権限なし→「管理者権限が必要です」\n・対象なし→エラー返却',
         '・ID採番はゼロ埋め5桁'),
        ('15.0', 'A001', 'パート別ミス数',
         '・クリアできなかった記録をユーザー×パートで集計し表形式で表示',
         '・なし', '・ミス数マトリクス',
         '・管理者権限が必要\n・管理者ユーザーは集計対象外',
         '・取得失敗→エラーメッセージ', '—'),
        ('16.0', '全画面', '認証切れの再ログイン誘導',
         '・APIが401を返したらセッションを破棄し、ログイン画面へ自動遷移\n'
         '・「セッションの有効期限が切れました。ログインし直してください。」を表示',
         '・APIレスポンス(401)', '・画面遷移+案内表示',
         '・全API呼び出し(ゲーム/選択/ランキング/管理/TTS)が対象',
         '・二重遷移は抑止', '要望No139/140対応'),
    ], wrap_from=4)

    # ---------------- Function Specs ----------------
    ws = new_sheet(wb, 'Function Specs', '機能仕様詳細（Function Specs）',
                   [('機能ID', 7), ('機能詳細ID', 10), ('機能名', 20), ('処理概要', 44),
                    ('入力', 22), ('出力', 22), ('条件・制約', 34), ('エラー処理', 32), ('備考', 26)])
    fill_rows(ws, [
        ('1', '1-1', 'パスワード認証',
         'usersシートからユーザー行を取得し、bcrypt.compareで照合(平文保存の場合は文字列比較)',
         'ユーザーID, パスワード', '認証結果(OK/NG)',
         '・usersシートA1:K/ヘッダー検証あり\n・user_idは先頭ゼロを保持(文字列扱い)',
         '不一致/未登録→401「認証に失敗しました」', 'bcrypt SALT_ROUNDS=10'),
        ('1', '1-2', 'JWT発行・Cookie設定',
         '認証成功時にJWT(userId/進捗/is_admin)を発行し、httpOnly Cookie authTokenに設定',
         '認証結果', 'Set-Cookie',
         '・有効期限24h\n・本番: secure+SameSite=None / 開発: SameSite=Lax',
         'JWT_SECRET未設定(本番)→起動エラー', 'iOS SafariはVercel rewriteで同一オリジン化'),
        ('2', '2-1', 'ウォームアップ(サーバー)',
         '起動時にGoogleクライアント初期化とusersシートの先読みキャッシュ(60秒)を実施',
         'なし', 'キャッシュ済みユーザーデータ',
         'Redis未接続でも起動は継続', '失敗時は警告ログのみ', '初回ログイン高速化'),
        ('2', '2-2', 'ウォームアップ(クライアント)',
         'ログイン画面マウント時に GET /health を送信(応答は破棄)',
         'なし', 'なし', 'ログイン中5秒経過で「サーバーを起動しています…」表示',
         '失敗は無視', '要望No136'),
        ('3', '3-1', 'ステージ一覧取得',
         'partsシートから、進捗(学年<現学年、同学年はパート<現パート、同パートはサブパート≦現在)以下を階層構造で返す',
         'ユーザーID', '{学年:{パート:[サブパート]}}+現在進捗',
         'JWTのuserIdと一致必須', '未登録→404', 'partsはA1:E/UNFORMATTED'),
        ('3', '3-2', '組み合わせ検証',
         '選択された学年/パート/サブパートがpartsシートに存在するか検証',
         '学年,パート,サブパート', 'valid(true/false)', '—', '不存在→「無効な組み合わせです」', '—'),
        ('4', '4-1', 'パート情報取得',
         'partsシートから該当行のpart_idとrequirement(回答条件)を返す',
         '学年,パート,サブパート', 'part_id, requirement',
         'キャッシュ10分', '不存在→404「該当partが見つかりません」', '—'),
        ('4', '4-2', '問題取得',
         'questionsシートをpart_idで絞り、display_order昇順で最大8問返す。answer_patternsを結合',
         'part_id', '問題配列(answers[]付き)',
         'MAX_QUESTIONS=8\nis_demoはtrue/false', '取得失敗→500', 'キャッシュ10分'),
        ('4', '4-3', 'デモ問題進行',
         '1問目(is_demo)は読み上げ2回→2秒後に自動で攻撃演出と正解提示(ユーザー入力なし)',
         '問題', '演出+正解表示', 'タイマー・正解カウント対象外', '—', '要望No89/92'),
        ('5', '5-1', 'TTS合成(サーバー)',
         'テキストをSSML化(エスケープ+phoneme発音補正+冒頭<break 200ms>)しGoogle Cloud TTSで合成',
         'text(≦1000字)', 'MP3(base64)',
         '・voice=en-US-Neural2-D, rate=0.95\n・Redisキャッシュ24h(キーに版数TTS_CACHE_VERSION)',
         '合成失敗→500(TTS-001)', '発音補正: read→/riːd/'),
        ('5', '5-2', 'TTS再生(クライアント)',
         '音声Blobを取得(セッションキャッシュ→API)し、読み込み完了を待ってから再生',
         'text', '再生完了(bool)',
         '・合成タイムアウト10秒→1回リトライ\n・読込待ち上限1.5秒\n・再生上限15秒',
         '失敗→console.errorに記録しfalse返却(ゲームは続行)', '要望No143/145/146'),
        ('5', '5-3', '読み上げシーケンス',
         '問題文を2回読み上げ(間隔1.2秒)→回答受付開始。正解時は正解文を1回読み上げ',
         '問題文/正解文', '音声再生',
         'マイクON中の問題読み上げはミュート', '—', '「回答条件」は音声で読まない'),
        ('5', '5-4', '先読み(プリフェッチ)',
         '問題間(intermission)に次問題の問題文・正解音声と画像を先読み',
         '次問題', 'キャッシュ充填', 'セッションキャッシュ上限50件', '失敗は無視', '要望No145/146'),
        ('6', '6-1', '音声認識制御',
         '銃ボタンでSpeechRecognition開始/停止。停止時に蓄積した認識候補で判定',
         'マイク入力', '認識テキスト群',
         'lang=en-US, continuous, interimResults, maxAlternatives=3',
         '認識エラーはログのみ(listening中は自動再開)', '認識なしで停止→判定せず継続'),
        ('7', '7-1', 'テキスト正規化',
         '小文字化→短縮形展開(I\'m→i am等)→記号除去→アポストロフィ落ち短縮形展開(im/dont等)→空白圧縮',
         '任意テキスト', '正規化済みテキスト',
         'ill/well/were/its等の実在語は誤変換防止のため展開しない',
         '—', '要望No149。正解・認識の両方に適用'),
        ('7', '7-2', '完全一致判定', '正規化済みの認識候補と正解パターンを文字列比較',
         '正規化済みテキスト', 'true/false', '—', '—', '—'),
        ('7', '7-3', 'トークン一致判定(1語正解)',
         '正解が1語のとき、認識文の単語列にその語が含まれていれば正解',
         '正規化済みテキスト', 'true/false', '正解が空白を含む場合は適用しない', '—',
         '要望No149(主語を答える問題対策)'),
        ('7', '7-4', '類似度判定',
         'レーベンシュタイン類似度またはJaccard係数(単語集合)が0.62以上なら正解',
         '正規化済みテキスト', 'true/false',
         'FUZZY_MATCH_THRESHOLD=0.62(変更時は要相談)', '—', '—'),
        ('8', '8-1', 'タイマー制御',
         '回答受付開始で30秒カウントダウン。残10秒で赤表示。正解/停止で解除',
         '—', '残り時間表示', 'TIME_LIMIT=30秒/デモは対象外',
         '時間切れ→タイマー停止→正解提示→次へ', '要望No80/83'),
        ('8', '8-2', '正解演出',
         'ビーム(0.8秒)→敵爆発(1秒)→正解表示(緑パネル+✓)→正解読み上げ→2.5秒のintermission',
         '判定=true', '演出+表示', '効果音attack.mp3', '—', '—'),
        ('8', '8-3', '不正解演出',
         '敵攻撃演出+miss.mp3。0.6秒後に再び回答受付(時間内)',
         '判定=false', '演出', '—', '—', '—'),
        ('9', '9-1', 'フリーズ検知',
         '5秒間隔でアクティビティを確認し、20秒無応答なら回復ダイアログを表示',
         '—', '回復ダイアログ',
         'FREEZE_TIMEOUT_MS=20000/FREEZE_CHECK_INTERVAL_MS=5000',
         '—', '要望No134/137'),
        ('9', '9-2', 'フリーズ回復',
         'リトライ=同一問題を再開/次の問題へ=スキップ/やめる=ステージ選択へ。音声・認識・タイマーは全停止',
         'ボタン押下', '状態復帰', '—', '—', '—'),
        ('11', '11-1', 'スコア保存',
         'scoresシートにscore_id自動採番で1行追記(userId/part_id/正解数/クリア/日時)',
         'スコア情報', 'score_id',
         'JWTのuserIdと一致必須/scoresは0〜1000', '書込失敗→500', '—'),
        ('12', '12-1', '進行判定',
         'クリア、または当該パートのプレイ回数≧10で次パートへ。partsを(学年,パート,サブパート)順に走査',
         'ユーザーID/現在地/クリア', '進行結果',
         '進捗の現在地とリクエストが一致すること', '不一致→progress mismatch',
         '最終パート→last part reached'),
        ('12', '12-2', '進捗更新',
         'usersシートのF〜H列(学年/パート/サブパート)とK列(更新日時)を更新',
         '次パート', '更新結果', '—', '更新失敗→500', 'クライアントの進捗表示も同期'),
        ('13', '13-1', 'ランキング集計',
         '当月のscoresをユーザー別に集計。挑戦=件数降順/正答率=平均スコア降順(同率はプレイ数→名前)',
         '当月スコア履歴', '上位3名×2種',
         '月キー(YYYY/MM)で60秒キャッシュ', '取得失敗→エラー表示', '—'),
        ('14', '14-1', 'ユーザー登録',
         '最終ID+1で採番(5桁ゼロ埋め)。8桁パスワード(英大小+数字を保証)を生成しbcryptで保存',
         'ニックネーム/本名', 'ユーザーID/初期パスワード',
         '管理者のみ/初期進捗は1-1-1', '書込失敗→500', '—'),
        ('14', '14-2', '進捗編集/PWリセット',
         '指定ユーザーの進捗(F〜H)更新、または新パスワード生成・ハッシュ保存',
         'ユーザーID+値', '更新結果/新パスワード', '管理者のみ', '対象なし→404', '—'),
        ('16', '16-1', '401検知',
         'apiFetch共通処理で401を検知したらセッション(localStorage)を破棄し/logInへ遷移',
         'APIレスポンス', '画面遷移+案内',
         'ログインAPI自体の401(認証失敗)は対象外', '—', '要望No139/140'),
    ], wrap_from=4)

    # ---------------- UI Specs ----------------
    ws = new_sheet(wb, 'UI Specs', '画面仕様（UI Specs）',
                   [('画面ID', 8), ('画面名', 14), ('表示要素', 44), ('操作内容', 32),
                    ('遷移先', 26), ('備考', 34)])
    fill_rows(ws, [
        ('G001', 'ログイン画面',
         '・タイトル「Welcome to English Game!」\n・User ID欄/Password欄\n・LOGINボタン'
         '(押下中は「Logging in...」+無効化)\n・5秒経過で「サーバーを起動しています…」\n'
         '・認証切れで戻された場合の案内メッセージ',
         '・ID/パスワードを入力\n・LOGIN押下',
         '・一般ユーザー→G002\n・管理者→A001',
         '・表示時に/healthへウォームアップping\n・失敗時は「メッセージ(エラーコード)」を表示'),
        ('G002', 'ステージ選択画面',
         '・タイトル「Select a Stage」\n・Ranking 🏆ボタン\n・学年/パート/サブパートのプルダウン'
         '(進捗以下のみ)\n・Game Startボタン',
         '・ステージを選択\n・Game Start押下',
         '・G003 ゲーム画面\n・G005 ランキング画面',
         '・開始時に無音を再生し自動再生制限を解除\n・現在進捗を初期選択'),
        ('G003', 'ゲーム画面',
         '・左上: Grade/Part/Subpart表示+「やめる」ボタン\n・右上: マイク状態(🎤MIC:ON/🔇MIC:OFF)'
         '+「Heard: 認識文」\n・左: 問題番号/残り時間(残10秒で赤)\n・上中央: 敵キャラ(通常/攻撃/撃破)\n'
         '・開始前: Requirement(回答条件)パネル+Startボタン\n・問題文(画像つき問題は左カラム表示)\n'
         '・画像(右カラム 最大300×250px)\n・銃ボタン(マイクON/OFF)\n・正解表示パネル(緑+✓)\n'
         '・フリーズ回復ダイアログ(リトライ/次の問題へ/やめる)',
         '・Startで開始\n・銃ボタンで回答(マイク)\n・やめるで中断',
         '・最終問題後→G004 結果画面\n・やめる→G002',
         '・問題文テキストはバナー表示後すぐ表示し、音声を2回読み上げてから回答受付\n'
         '・バナー「start a demo !」「Question N !」を2秒表示\n・デモ問題は自動進行'),
        ('G004', '結果画面',
         '・スコア(正解数/問題数と%)\n・CLEAR/FAILEDバッジ\n・メッセージ(Perfect!〜Try Again...)\n'
         '・クリア時「次のステージが解放されました！」\n・未クリア時「あと N 問正解でクリア！」\n'
         '・LOGOUTボタン/NEXT(またはRetry)ボタン',
         '・NEXT/Retry→ステージ選択へ\n・LOGOUT→ログインへ',
         '・G002 / G001',
         '・クリア条件: 5問正解(デモ除く7問中)\n・残り数はCORRECT_TO_CLEAR基準(要望No133)'),
        ('G005', 'ランキング画面',
         '・タイトル「Ranking 👑」\n・期間(当月)表示\n・Number of try(挑戦回数 上位3)\n'
         '・Best Scores(正答率 上位3)\n・Backボタン',
         '・Back押下', '・G002',
         '・1位金/2位銀/3位銅の装飾\n・データなしは空欄表示'),
        ('A001', '管理者画面',
         '・新規登録フォーム(名前/ニックネーム)→発行ID・初期PW表示\n・パスワードリセット(ユーザーID指定)\n'
         '・ユーザー一覧(進捗の編集/保存/キャンセル)\n・パート別ミス数マトリクス',
         '・登録/リセット/編集操作',
         '・(遷移なし)',
         '・管理者アカウントでログインした場合のみ\n・一般ユーザーは一覧から除外'),
    ], wrap_from=3)

    # ---------------- Data Specs ----------------
    ws = new_sheet(wb, 'Data Specs', 'データ仕様（Data Specs）',
                   [('シート名(DB)', 16), ('主キー', 12), ('カラム名', 16), ('データ形式', 14),
                    ('説明', 40), ('備考', 36)])
    rows = []

    def table(name, pk, cols):
        for i, (col, typ, desc, note) in enumerate(cols):
            rows.append((name if i == 0 else '', pk if i == 0 else '', col, typ, desc, note))

    table('users', 'id', [
        ('id', 'NUMBER', 'ユーザー内部ID(連番)', 'PK'),
        ('user_id', 'TEXT(5桁)', 'ログイン用ID(例: 00001)', 'ゼロ埋め。文字列として保持(FORMATTED_VALUE)'),
        ('password', 'TEXT', 'bcryptハッシュ($2b$…)', '平文が残っている場合も照合可(後方互換)'),
        ('nickname', 'TEXT', 'ニックネーム', 'ランキング表示名'),
        ('real_name', 'TEXT', '本名', '管理画面表示'),
        ('current_grade', 'NUMBER', '現在の学年(1-3)', '管理画面で編集可'),
        ('current_part', 'NUMBER', '現在のパート', '〃'),
        ('current_subpart', 'NUMBER', '現在のサブパート', '〃'),
        ('is_admin', 'BOOLEAN', '管理者権限', 'true/false'),
        ('created_at', 'TEXT', '作成日時(YYYY/MM/DD HH:MM:SS)', ''),
        ('updated_at', 'TEXT', '更新日時', '進行時に自動更新'),
    ])
    table('parts', 'part_id', [
        ('part_id', 'NUMBER', 'パートID', 'PK。学年×1000+パート×10+サブパート(例: 3-8-2→3082)'),
        ('grade_id', 'NUMBER', '学年(1-3)', ''),
        ('part_no', 'NUMBER', 'パート番号', ''),
        ('subpart_no', 'NUMBER', 'サブパート番号', ''),
        ('requirement', 'TEXT', '回答条件(ゲーム開始前に表示)', '改行可。音声では読まない'),
    ])
    table('questions', 'question_id', [
        ('question_id', 'NUMBER', '問題ID', 'PK。part_id×10+出題順(例: 3082の7問目→30827)'),
        ('part_id', 'NUMBER', '紐づくパートID', 'FK→parts'),
        ('display_order', 'NUMBER', '出題順(1-8)', 'パート内で一意'),
        ('is_demo', 'BOOLEAN', 'デモ問題フラグ', '各パートの1問目のみtrue'),
        ('question_text', 'TEXT', '問題文(読み上げ・表示)', '1-44-1のみ問題文なし(イラスト出題)'),
        ('image_url', 'TEXT', '画像URL(/questions/形式)', '命名: 学年_パート_サブパート_番号(2桁).png'),
    ])
    table('answer_patterns', 'id', [
        ('id', 'NUMBER', '解答パターンID(連番)', 'PK'),
        ('question_id', 'NUMBER', '紐づく問題ID', 'FK→questions。複数パターン可'),
        ('expected_text', 'TEXT', '期待解答', "音声認識対策の別解を含む(例: 'I'に対する'eye')"),
    ])
    table('scores', 'score_id', [
        ('score_id', 'NUMBER', 'スコアID(自動採番)', 'PK'),
        ('user_id', 'TEXT', 'プレイしたユーザーID', 'FK→users.user_id'),
        ('part_id', 'NUMBER', 'プレイしたパートID', 'FK→parts'),
        ('scores', 'NUMBER', '正解数(デモ除く)', '0〜1000で検証'),
        ('clear', 'BOOLEAN', 'クリア(5問以上正解)', ''),
        ('play_date', 'TEXT', 'プレイ日時(YYYY/MM/DD HH:MM:SS)', 'ランキングの月判定に使用'),
    ])
    rows.append(('(共通)', '', '', '', 'ヘッダー行はAPI側で毎回検証(不一致は500)',
                 '列の追加・並び替えはシステム改修が必要'))
    rows.append(('(規模)', '', '', '', 'parts 276 / questions 2,208 / answer_patterns 2,211(2026-07-02時点)',
                 '再構築手順: scripts/build_from_question_sheet.py'))
    fill_rows(ws, rows, wrap_from=5)

    # ---------------- Tech Stack ----------------
    ws = new_sheet(wb, 'Tech Stack', '使用技術まとめ（Tech Stack）',
                   [('項目', 18), ('技術・ライブラリ名', 34), ('説明', 70)])
    fill_rows(ws, [
        ('フロントエンド', 'React 19 / TypeScript 5.8 / Vite 6', 'SPA。React Router 7で画面遷移'),
        ('バックエンド', 'Express 5 (Node.js, CommonJS)', 'REST API。入力検証・エラーコード付きレスポンス'),
        ('DB', 'Google スプレッドシート', 'users/parts/questions/answer_patterns/scores の5シート'),
        ('キャッシュ', 'Redis', 'シート10分/TTS音声24時間(版数つき)/ランキング60秒。未接続でも動作'),
        ('音声出力', 'Google Cloud Text-to-Speech', 'en-US-Neural2-D, rate 0.95。SSML(発音補正/冒頭200ms無音)'),
        ('音声認識', 'Web Speech API', 'ブラウザ内蔵。lang=en-US, continuous+interim, 候補3'),
        ('認証', 'JWT + httpOnly Cookie / bcrypt', 'authToken(24時間)。パスワードはbcryptハッシュ'),
        ('デプロイ(フロント)', 'Vercel', '/api/* をRenderへrewriteし同一オリジン化(iOS Safari ITP対策)'),
        ('デプロイ(バック)', 'Render', '無料プランはコールドスタートあり→起動時ウォームアップ+画面案内'),
        ('検証環境', 'DATA_SOURCE=local モード', 'Google認証なしで backend/data/*.json を使いローカル動作確認'),
        ('データ整備', 'Python 3 / openpyxl', 'questionシート→DB再構築・検証・カタログ生成(scripts/)'),
        ('フォント', 'Playpen Sans', 'ゲーム内テキスト(要望No65)'),
    ], wrap_from=3)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f'生成完了: {OUT}')


if __name__ == '__main__':
    main()
