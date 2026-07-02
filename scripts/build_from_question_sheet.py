#!/usr/bin/env python3
"""questionシート(最終ソース)から DB データ一式を再構築する。

入力(読み取りのみ):
    QUESTION_XLSX 環境変数 or リポジトリ親ディレクトリの
    「English Speaking Drill仕様書.xlsx」の question シート … 問題データの正
    outputs/Data Specs_2026-06-11.xlsx … users/scores の引き継ぎ元(本番状態)

出力:
    backend/data/{users,grades,parts,questions,answer_patterns,scores}.json
        … DATA_SOURCE=local 用シード
    outputs/DB_<日付>.xlsx … Google Sheets へ反映するDB完成版
    scripts/reports/rebuild_report.md … 解析ログ・異常値・突合結果

正規化ルール(詳細は rebuild_report.md にも記録):
    - 学年ラベル: 中１英語→1 / 2学年→2 / 中３→3。空白のみセルは前値継続
    - パート/サブパートも空白のみセルは前値継続。学年切替時はリセットし、
      欠落していれば 1 で開始(レポートに記録)
    - ID体系: part_id = 学年*1000 + パート*10 + サブパート
              question_id = part_id*10 + 出題順
    - デモ: 各パートの1問目(出題順1)を機械的に true(F列の〇は参考値)
    - 英文: スマートクォート→直線、全角空白→半角、連続空白圧縮、前後trim
    - 回答条件(日本語): 前後trimのみ(改行・全角は保持)
    - 日本語訳: 通常 L/M 列。中3後半は K/L 列にずれているため自動判定
    - 画像: J列を正規化(空白除去・2桁ゼロ埋め・'0.6png'等の誤記修正)。
      3-8-2 のみJ列が空のため命名規則から補完(要望No144)
    - 解答パターン: I列を主とし、解答が 'I' 1語の問題には音声認識対策の
      別解 'eye' を追加(要望No149)

RECON_DIR 環境変数に旧データセット(JSON群)のディレクトリを指定すると、
再構築結果との差分を突合してレポートに出力する。
"""
import json
import os
import re
import sys
import unicodedata
from datetime import date

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), 'lib'))
from ids import part_id as make_part_id, question_id as make_question_id  # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
QUESTION_XLSX = os.environ.get(
    'QUESTION_XLSX',
    os.path.join(os.path.dirname(ROOT), 'English Speaking Drill仕様書.xlsx'),
)
PREV_EXPORT = os.path.join(ROOT, 'outputs', 'Data Specs_2026-06-11.xlsx')
DATA_DIR = os.path.join(ROOT, 'backend', 'data')
IMG_DIR = os.path.join(ROOT, 'frontend', 'public', 'questions')
REPORT = os.path.join(ROOT, 'scripts', 'reports', 'rebuild_report.md')
RECON_DIR = os.environ.get('RECON_DIR')

TODAY = os.environ.get('BUILD_DATE') or date.today().isoformat()
DB_XLSX = os.path.join(ROOT, 'outputs', f'DB_{TODAY}.xlsx')

GRADE_LABELS = {'中１英語': 1, '2学年': 2, '中３': 3}
CJK_RE = re.compile(r'[぀-ヿ一-鿿]')

# J列が空でも命名規則から画像を補完するパート(要望No144: 3-8-2追加分)
DERIVE_IMAGE_PARTS = {3082}

# 音声認識対策の別解(要望No149): 主解答(正規化後) → 追加パターン
ANSWER_ALIASES = {
    'I': ['eye'],
}


def blank(v):
    return v is None or str(v).replace('　', ' ').strip() == ''


def clean_en(v):
    """英文の正規化: スマートクォート→直線 / 全角空白→半角 / 連続空白圧縮 / trim"""
    if v is None:
        return ''
    s = str(v)
    for a, b in (('’', "'"), ('‘', "'"), ('“', '"'), ('”', '"'), ('　', ' '), (' ', ' '),
                 ('？', '?'), ('！', '!'), ('，', ','), ('．', '.')):
        s = s.replace(a, b)
    s = re.sub(r'[ \t]+', ' ', s)
    return s.strip()


def clean_jp(v):
    """日本語テキスト: 前後の空白だけ落とす(改行・全角空白は保持)"""
    if v is None:
        return ''
    return str(v).strip()


def normalize_image_name(raw):
    """J列のファイル名を正規形 {g}_{p}_{s}_{NN}.png に直す。不正は None。"""
    if blank(raw):
        return None
    base = unicodedata.normalize('NFKC', str(raw))
    base = base.replace(' ', '').replace('　', '')
    base = re.sub(r'\.?png$', '', base, flags=re.I)
    base = base.replace('-', '_')
    m = re.fullmatch(r'(\d+)_(\d+)_(\d+)_(\d+(?:\.\d+)?)', base)
    if not m:
        return None
    g, p, s, n = m.groups()
    n = n.replace('.', '')  # '0.6' のような打ち間違い → '06'
    return f'{int(g)}_{int(p)}_{int(s)}_{int(n):02d}.png'


def to_int(v):
    return int(float(v))


def parse_question_sheet(report):
    wb = openpyxl.load_workbook(QUESTION_XLSX, data_only=True)
    ws = wb['question']

    parts = {}       # part_id -> dict
    questions = []   # dict list
    anomalies = report['anomalies']

    grade = None
    part_no = None
    subpart_no = None

    for i, r in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        # 列: A学年 B章 Cパート Dサブパート E問題番号 Fデモ G回答条件 H問題 I解答
        #     Jファイル名 K/L/M 日本語訳(位置ずれあり) Nメモ O〜 作業メモ
        g_raw, _ch, pt_raw, sp_raw, no_raw = r[0], r[1], r[2], r[3], r[4]
        demo_raw, req_raw, q_raw, a_raw, img_raw = r[5], r[6], r[7], r[8], r[9]

        if not blank(g_raw):
            label = str(g_raw).strip()
            if label in GRADE_LABELS:
                new_grade = GRADE_LABELS[label]
                if new_grade != grade:
                    grade = new_grade
                    part_no = None
                    subpart_no = None
            else:
                anomalies.append(f'row{i}: 未知の学年ラベル {label!r} → 前値({grade})を継続')
        elif g_raw is not None and str(g_raw) != '' and grade is not None:
            # 全角スペースのみのセル(例: 中3の途中) → 前値継続
            pass

        if not blank(pt_raw):
            part_no = to_int(pt_raw)
        if not blank(sp_raw):
            subpart_no = to_int(sp_raw)

        if blank(no_raw):
            continue  # 問題行ではない

        if grade is None:
            anomalies.append(f'row{i}: 学年不明の問題行をスキップ')
            continue
        if part_no is None:
            part_no = 1
            anomalies.append(f'row{i}: パート欠落 → 1 で開始 (学年{grade})')
        if subpart_no is None:
            subpart_no = 1
            anomalies.append(f'row{i}: サブパート欠落 → 1 で開始 (学年{grade}-{part_no})')

        pid = make_part_id(grade, part_no, subpart_no)
        order = to_int(no_raw)
        qid = make_question_id(grade, part_no, subpart_no, order)

        if pid not in parts:
            parts[pid] = {
                'part_id': pid, 'grade_id': grade, 'part_no': part_no,
                'subpart_no': subpart_no, 'requirement': '',
            }
        if not blank(req_raw) and not parts[pid]['requirement']:
            parts[pid]['requirement'] = clean_jp(req_raw)

        # 日本語訳: 通常 L/M。中3後半は K/L にずれる(K列に日本語がありLがその続き)
        k, l, m = r[10], r[11], r[12]
        if not blank(l) or not blank(m):
            jp_q, jp_a = clean_jp(l), clean_jp(m)
        elif not blank(k) and CJK_RE.search(str(k)):
            jp_q, jp_a = clean_jp(k), clean_jp(l)
        else:
            jp_q, jp_a = '', ''
        # K列にも日本語がある通常行(KLM)は K を備考へ回す
        note_extras = []
        if not blank(k) and (not blank(l) or not blank(m)) and CJK_RE.search(str(k)):
            note_extras.append(clean_jp(k))

        # 画像
        img_name = normalize_image_name(img_raw)
        if not blank(img_raw) and img_name is None:
            anomalies.append(f'row{i}: ファイル名を正規化できない {str(img_raw)!r} (問題 {qid})')
        if img_name is None and pid in DERIVE_IMAGE_PARTS:
            img_name = f'{grade}_{part_no}_{subpart_no}_{order:02d}.png'
            report['derived_images'].append(f'{qid}: {img_name}')
        if img_name is not None and str(img_raw or '').strip() != img_name:
            report['fixed_image_names'].append(f'{qid}: {str(img_raw)!r} → {img_name}')

        # 作業メモ(N〜T列) — DBには入れずカタログの備考として保持
        notes = [clean_jp(x) for x in [r[13], *r[14:20]] if not blank(x)]
        notes = [n for n in notes if n not in ('True', 'False')]
        notes.extend(note_extras)

        questions.append({
            'question_id': qid,
            'part_id': pid,
            'display_order': order,
            'is_demo': order == 1,
            'question_text': clean_en(q_raw),
            'image_url': f'/questions/{img_name}' if img_name else '',
            'answer': clean_en(a_raw),
            'demo_flag_in_sheet': not blank(demo_raw),
            'jp_question': jp_q,
            'jp_answer': jp_a,
            'notes': ' / '.join(notes),
            'row': i,
        })

    return parts, questions


def build_answer_patterns(questions, report):
    patterns = []
    next_id = 1
    for q in questions:
        ans = q['answer']
        if not ans:
            report['anomalies'].append(f"question {q['question_id']}: 解答が空")
            continue
        patterns.append({'id': next_id, 'question_id': q['question_id'], 'expected_text': ans})
        next_id += 1
        for alias in ANSWER_ALIASES.get(ans, []):
            patterns.append({'id': next_id, 'question_id': q['question_id'], 'expected_text': alias})
            report['alias_patterns'].append(f"{q['question_id']}: {ans!r} に別解 {alias!r} を追加")
            next_id += 1
    return patterns


def fmt_ts(v):
    """日時セルを本番形式 'YYYY/MM/DD HH:MM:SS' の文字列にする"""
    if v is None:
        return ''
    if hasattr(v, 'strftime'):
        return v.strftime('%Y/%m/%d %H:%M:%S')
    return str(v).replace('-', '/')


def carry_over_users_scores():
    wb = openpyxl.load_workbook(PREV_EXPORT, data_only=True)
    users = []
    for r in wb['users'].iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        users.append({
            'id': int(r[0]), 'user_id': str(r[1]), 'password': str(r[2]),
            'nickname': str(r[3] or ''), 'real_name': str(r[4] or ''),
            'current_grade': int(r[5]), 'current_part': int(r[6]), 'current_subpart': int(r[7]),
            'is_admin': bool(r[8]), 'created_at': fmt_ts(r[9]), 'updated_at': fmt_ts(r[10]),
        })
    scores = []
    for r in wb['scores'].iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        scores.append({
            'score_id': int(r[0]), 'user_id': str(r[1]), 'part_id': int(r[2]),
            'scores': int(r[3]), 'clear': bool(r[4]), 'play_date': fmt_ts(r[5]),
        })
    return users, scores


def validate(parts, questions, patterns, report):
    errors = report['errors']
    # 8問/パート・出題順1..8一意
    by_part = {}
    for q in questions:
        by_part.setdefault(q['part_id'], []).append(q)
    for pid, qs in sorted(by_part.items()):
        orders = sorted(x['display_order'] for x in qs)
        if orders != list(range(1, 9)):
            errors.append(f'part {pid}: 出題順が1〜8でない: {orders}')
    # question_id 一意
    qids = [q['question_id'] for q in questions]
    dup = {x for x in qids if qids.count(x) > 1}
    if dup:
        errors.append(f'question_id 重複: {sorted(dup)}')
    # 画像実在
    for q in questions:
        if q['image_url']:
            fn = q['image_url'].split('/')[-1]
            if not os.path.exists(os.path.join(IMG_DIR, fn)):
                errors.append(f"question {q['question_id']}: 画像なし {fn}")
    # requirement
    for pid, p in sorted(parts.items()):
        if not p['requirement']:
            report['anomalies'].append(f'part {pid}: 回答条件(requirement)が空')
    # デモ列との突合(参考)
    for q in questions:
        if q['is_demo'] != q['demo_flag_in_sheet']:
            report['demo_mismatch'].append(
                f"{q['question_id']}: シートのデモ〇={'あり' if q['demo_flag_in_sheet'] else 'なし'}"
                f" / 構造ルール(1問目)={'デモ' if q['is_demo'] else '通常'}")


def reconcile(parts, questions, patterns, report):
    if not RECON_DIR or not os.path.isdir(RECON_DIR):
        return
    with open(os.path.join(RECON_DIR, 'questions.json'), encoding='utf-8') as f:
        old_q = {q['question_id']: q for q in json.load(f)}
    with open(os.path.join(RECON_DIR, 'parts.json'), encoding='utf-8') as f:
        old_p = {p['part_id']: p for p in json.load(f)}
    with open(os.path.join(RECON_DIR, 'answer_patterns.json'), encoding='utf-8') as f:
        old_a = {}
        for a in json.load(f):
            old_a.setdefault(a['question_id'], []).append(a['expected_text'])

    new_q = {q['question_id']: q for q in questions}
    new_p = parts
    r = report['recon']
    r['parts_only_old'] = sorted(set(old_p) - set(new_p))
    r['parts_only_new'] = sorted(set(new_p) - set(old_p))
    r['q_only_old'] = sorted(set(old_q) - set(new_q))
    r['q_only_new'] = sorted(set(new_q) - set(old_q))
    text_diff = []
    ans_diff = []
    img_diff = []
    new_a = {}
    for p in patterns:
        new_a.setdefault(p['question_id'], []).append(p['expected_text'])
    for qid in sorted(set(old_q) & set(new_q)):
        o, n = old_q[qid], new_q[qid]
        if clean_en(o['question_text']) != n['question_text']:
            text_diff.append((qid, o['question_text'], n['question_text']))
        if str(o.get('image_url') or '') != n['image_url']:
            img_diff.append((qid, o.get('image_url') or '', n['image_url']))
        oa = [clean_en(x) for x in old_a.get(qid, [])]
        na = new_a.get(qid, [])
        if oa and na and clean_en(oa[0]) != na[0]:
            ans_diff.append((qid, oa[0], na[0]))
    r['text_diff'] = text_diff
    r['answer_diff'] = ans_diff
    r['image_diff'] = img_diff
    for pid in sorted(set(old_p) & set(new_p)):
        if clean_jp(old_p[pid].get('requirement') or '') != new_p[pid]['requirement']:
            r.setdefault('req_diff', []).append(pid)


def write_json(parts, questions, patterns, users, scores):
    os.makedirs(DATA_DIR, exist_ok=True)
    grades = [{'id': g, 'grade_no': g} for g in (1, 2, 3)]
    q_out = [
        {k: q[k] for k in ('question_id', 'part_id', 'display_order', 'is_demo', 'question_text', 'image_url')}
        for q in questions
    ]
    files = {
        'grades': grades,
        'parts': [parts[k] for k in sorted(parts)],
        'questions': q_out,
        'answer_patterns': patterns,
        'users': users,
        'scores': scores,
    }
    for name, data in files.items():
        with open(os.path.join(DATA_DIR, f'{name}.json'), 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
            f.write('\n')
    # 問題カタログ生成用の拡張データ(日本語訳・シート内メモつき)
    catalog = [
        {k: q[k] for k in ('question_id', 'part_id', 'display_order', 'is_demo',
                           'question_text', 'image_url', 'answer',
                           'jp_question', 'jp_answer', 'notes')}
        for q in questions
    ]
    os.makedirs(os.path.join(ROOT, 'scripts', 'reports'), exist_ok=True)
    with open(os.path.join(ROOT, 'scripts', 'reports', 'catalog_source.json'), 'w', encoding='utf-8') as f:
        json.dump(catalog, f, ensure_ascii=False, indent=1)
        f.write('\n')


def write_db_xlsx(parts, questions, patterns, users, scores):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    wb.remove(wb.active)
    head_font = Font(bold=True, color='FFFFFF')
    head_fill = PatternFill('solid', fgColor='4472C4')

    def add_sheet(name, header, rows):
        ws = wb.create_sheet(name)
        ws.append(header)
        for c in ws[1]:
            c.font = head_font
            c.fill = head_fill
        for row in rows:
            ws.append(row)
        ws.freeze_panes = 'A2'

    add_sheet('users',
              ['id', 'user_id', 'password', 'nickname', 'real_name', 'current_grade',
               'current_part', 'current_subpart', 'is_admin', 'created_at', 'updated_at'],
              [[u['id'], u['user_id'], u['password'], u['nickname'], u['real_name'],
                u['current_grade'], u['current_part'], u['current_subpart'], u['is_admin'],
                u['created_at'], u['updated_at']] for u in users])
    add_sheet('grades', ['id', 'grade_no'], [[g, g] for g in (1, 2, 3)])
    add_sheet('parts', ['part_id', 'grade_id', 'part_no', 'subpart_no', 'requirement'],
              [[p['part_id'], p['grade_id'], p['part_no'], p['subpart_no'], p['requirement']]
               for p in (parts[k] for k in sorted(parts))])
    add_sheet('questions',
              ['question_id', 'part_id', 'display_order', 'is_demo', 'question_text', 'image_url'],
              [[q['question_id'], q['part_id'], q['display_order'], q['is_demo'],
                q['question_text'], q['image_url']] for q in questions])
    add_sheet('answer_patterns', ['id', 'question_id', 'expected_text'],
              [[a['id'], a['question_id'], a['expected_text']] for a in patterns])
    add_sheet('scores', ['score_id', 'user_id', 'part_id', 'scores', 'clear', 'play_date'],
              [[s['score_id'], s['user_id'], s['part_id'], s['scores'], s['clear'], s['play_date']]
               for s in scores])
    wb.save(DB_XLSX)


def write_report(parts, questions, patterns, users, scores, report):
    os.makedirs(os.path.dirname(REPORT), exist_ok=True)
    lines = []
    lines.append('# データ再構築レポート')
    lines.append('')
    lines.append(f'- 実行日: {TODAY}')
    lines.append(f'- 入力: `{os.path.basename(QUESTION_XLSX)}` の question シート(最終ソース)')
    lines.append(f'- 出力: `backend/data/*.json`, `outputs/{os.path.basename(DB_XLSX)}`')
    lines.append('')
    g_counts = {}
    for p in parts.values():
        g_counts[p['grade_id']] = g_counts.get(p['grade_id'], 0) + 1
    lines.append('## 件数サマリ')
    lines.append('')
    lines.append(f"- パート: {len(parts)} (学年別: {dict(sorted(g_counts.items()))})")
    lines.append(f'- 問題: {len(questions)}')
    lines.append(f'- 解答パターン: {len(patterns)}')
    lines.append(f'- ユーザー: {len(users)} / スコア履歴: {len(scores)} (前回exportから引き継ぎ)')
    lines.append('')

    def section(title, items, fmt=lambda x: str(x)):
        lines.append(f'## {title} ({len(items)}件)')
        lines.append('')
        if not items:
            lines.append('なし')
        for it in items:
            lines.append(f'- {fmt(it)}')
        lines.append('')

    section('エラー', report['errors'])
    section('解析時の補正・異常値', report['anomalies'])
    section('画像ファイル名の補正', report['fixed_image_names'])
    section('命名規則から補完した画像 (3-8-2 / 要望No144)', report['derived_images'])
    section('音声認識対策の追加別解 (要望No149)', report['alias_patterns'])
    section('デモ〇列と構造ルールの不一致(参考・構造ルールを採用)', report['demo_mismatch'])

    r = report['recon']
    if r:
        lines.append('## 旧データセット(featureブランチ 2026-06-11版)との突合')
        lines.append('')
        lines.append('questionシートを正とし、差分は「Mukaさんの6/11以降の編集」または')
        lines.append('「旧版のみの修正(要Muka確認)」に分類する。')
        lines.append('')
        lines.append(f"- 旧のみに存在するパート: {r.get('parts_only_old', [])} ← **要Muka確認**")
        lines.append(f"- 新のみに存在するパート: {r.get('parts_only_new', [])}")
        lines.append(f"- 旧のみに存在する問題: {len(r.get('q_only_old', []))}件 {r.get('q_only_old', [])[:20]}")
        lines.append(f"- 新のみに存在する問題: {len(r.get('q_only_new', []))}件 {r.get('q_only_new', [])[:20]}")
        lines.append(f"- 問題文が異なる: {len(r.get('text_diff', []))}件")
        for qid, o, n in r.get('text_diff', [])[:50]:
            lines.append(f'    - {qid}: {o!r} → {n!r}')
        if len(r.get('text_diff', [])) > 50:
            lines.append(f"    - …ほか{len(r['text_diff']) - 50}件")
        lines.append(f"- 解答が異なる: {len(r.get('answer_diff', []))}件")
        for qid, o, n in r.get('answer_diff', [])[:50]:
            lines.append(f'    - {qid}: {o!r} → {n!r}')
        if len(r.get('answer_diff', [])) > 50:
            lines.append(f"    - …ほか{len(r['answer_diff']) - 50}件")
        lines.append(f"- 画像URLが異なる: {len(r.get('image_diff', []))}件")
        for qid, o, n in r.get('image_diff', [])[:30]:
            lines.append(f'    - {qid}: {o!r} → {n!r}')
        lines.append(f"- 回答条件が異なるパート: {r.get('req_diff', [])}")
        lines.append('')

    with open(REPORT, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines) + '\n')


def main():
    report = {
        'errors': [], 'anomalies': [], 'fixed_image_names': [], 'derived_images': [],
        'alias_patterns': [], 'demo_mismatch': [], 'recon': {},
    }
    parts, questions = parse_question_sheet(report)
    patterns = build_answer_patterns(questions, report)
    users, scores = carry_over_users_scores()
    validate(parts, questions, patterns, report)
    reconcile(parts, questions, patterns, report)
    write_report(parts, questions, patterns, users, scores, report)

    if report['errors']:
        print(f"ERROR: {len(report['errors'])}件 (詳細: {os.path.relpath(REPORT, ROOT)})")
        for e in report['errors'][:20]:
            print(f'  - {e}')
        return 1

    write_json(parts, questions, patterns, users, scores)
    write_db_xlsx(parts, questions, patterns, users, scores)
    g_counts = {}
    for p in parts.values():
        g_counts[p['grade_id']] = g_counts.get(p['grade_id'], 0) + 1
    print(f'パート: {len(parts)} {dict(sorted(g_counts.items()))} / 問題: {len(questions)} / '
          f'解答パターン: {len(patterns)}')
    print(f'OK: backend/data/*.json, {os.path.relpath(DB_XLSX, ROOT)}, '
          f'{os.path.relpath(REPORT, ROOT)}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
