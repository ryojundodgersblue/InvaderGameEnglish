#!/usr/bin/env python3
"""作業結果をExcelワークブックへ反映する(ドキュメント更新)。

更新対象:
    outputs/Mukaさん_ユーザー要望書.xlsx
        - コンテンツ修正依頼: apply_report.json の結果から 対応状況/担当 を更新
        - 不具合・改善報告: 今回対応した項目を完了化+対応内容を記入
        - 機能・仕様要望: No101(速さランキング)を完了化
    outputs/English Speaking Drill仕様書_実装整合版.xlsx
        - Revision: v1.6 行を追加
        - Data Specs: scores に avg_answer_time 行を追加
        - Function List / UI Specs: ランキング3種化を反映

セル値のみ変更し、書式・他シートには触れない。
"""
import json
import os
import sys

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REQ_XLSX = os.path.join(ROOT, 'outputs', 'Mukaさん_ユーザー要望書.xlsx')
SPEC_XLSX = os.path.join(ROOT, 'outputs', 'English Speaking Drill仕様書_実装整合版.xlsx')
REPORT_JSON = os.path.join(ROOT, 'scripts', 'reports', 'apply_report.json')
TODAY = '2026-06-11'

# 不具合・改善報告の対応内容(No → (状況, 対応内容))
BUG_RESOLUTIONS = {
    132: ('完了', '完成版データで1-24-1/1-24-2/1-25-1の問題文・解答を整備し、1-24-1/1-25-1のイラストを追加しました。音声はTTSで合成されます。※1-24-2のイラストは未提供のため画像なしで動作します(提供あり次第追加可能)'),
    133: ('完了', 'クリアまでの残り問題数の計算を修正しました(あと[5-正解数]問正解でクリア!と表示)。実機で確認済みです'),
    134: ('完了', '画面停止時の回復ダイアログを改善しました(リトライ/次の問題へ/やめる の3択)。検知時間も30秒→20秒に短縮し、生徒が考えている時間を停止と誤判定しないようにしました'),
    136: ('完了', 'ログイン画面を開いた時点でサーバーを事前起動するようにし、待ち時間を短縮しました。5秒以上かかる場合は「サーバー起動中です」の案内を表示します。※完全な解消にはサーバープランの変更が必要なため、納品時にご相談します'),
    137: ('完了', '画面停止の検知を20秒に短縮し、「次の問題へ」ボタンを含む回復ダイアログがすぐ表示されるようにしました'),
    138: ('完了', 'readを現在形(リード)で発音するよう修正しました。修正前に作られた音声キャッシュも無効化したため、確実に新しい発音になります'),
    141: ('完了', 'イラストがある問題では、問題文を左側・イラストを右側に分けて表示するレイアウトに変更し、どの画面サイズでも重ならないようにしました'),
    142: ('完了', 'No138と同じ対応です。readは現在形(リード)で発音されます'),
    143: ('完了', '音声の先頭に0.2秒の無音を追加し、再生前に読み込み完了を待つようにしたため、冒頭の音が欠けなくなりました'),
    146: ('完了', '音声の取得に失敗した場合に自動で再取得するようにし、2回読み上げが欠けないようにしました'),
    148: ('完了', '2026-05-12に対応済み・ご確認済みの件です。完成版データでも全パートが正しく出題されることを自動チェックで確認しました'),
    149: ('完了', '2026-05-28にMukaさんに再確認いただき問題ないことを確認済みです。今回の修正依頼反映で解答パターンも最新化しています'),
}

FEATURE_RESOLUTIONS = {
    101: ('完了', '正答率(Best Scores)と速さ(Fastest Answers: 回答までの平均秒)のランキングを実装しました。挑戦回数と合わせて3種類のランキングが表示されます'),
}


def update_requests_workbook():
    wb = openpyxl.load_workbook(REQ_XLSX)

    # --- コンテンツ修正依頼 ---
    with open(REPORT_JSON, encoding='utf-8') as f:
        report = json.load(f)
    # 行番号 → 結果(複数エントリは applied > superseded > already_ok > skipped の優先で集約)
    rank = {'applied': 3, 'superseded': 2, 'already_ok': 1, 'skipped': 0}
    row_result = {}
    for e in report['entries']:
        cur = row_result.get(e['row'])
        if cur is None or rank[e['action']] > rank[cur]:
            row_result[e['row']] = e['action']

    ws = wb['コンテンツ修正依頼']
    updated = {'完了': 0, '要確認': 0}
    for row, action in row_result.items():
        status_cell = ws.cell(row=row, column=5)
        owner_cell = ws.cell(row=row, column=6)
        if action == 'skipped':
            status_cell.value = '要確認'
            updated['要確認'] += 1
        else:
            status_cell.value = '完了'
            updated['完了'] += 1
        owner_cell.value = '藤岡'
    # 件数サマリ行(2行目)を更新
    total = sum(1 for r in ws.iter_rows(min_row=4, values_only=True) if r and (r[0] is not None or r[1] is not None))
    done = sum(1 for r in ws.iter_rows(min_row=4, values_only=True)
               if r and (r[0] is not None or r[1] is not None) and str(r[4] or '').strip() == '完了')
    ws.cell(row=2, column=1).value = (
        f'問題文・Requirement・イラストの修正依頼一覧。 未対応: {total - done}件  完了: {done}件'
    )
    print(f'コンテンツ修正依頼: 完了 {updated["完了"]}行 / 要確認 {updated["要確認"]}行 (総数{total}・完了{done})')

    # --- 不具合・改善報告 ---
    ws = wb['不具合・改善報告']
    count = 0
    for row in ws.iter_rows(min_row=4):
        no = row[0].value
        if no in BUG_RESOLUTIONS:
            status, note = BUG_RESOLUTIONS[no]
            row[4].value = note      # 対応内容
            row[5].value = status    # 対応状況
            if not row[6].value:
                row[6].value = TODAY  # 完了日
            count += 1
    print(f'不具合・改善報告: {count}件を完了化')

    # --- 機能・仕様要望 ---
    ws = wb['機能・仕様要望']
    count = 0
    for row in ws.iter_rows(min_row=4):
        no = row[0].value
        if no in FEATURE_RESOLUTIONS:
            status, note = FEATURE_RESOLUTIONS[no]
            row[4].value = note      # 回答・決定事項
            row[5].value = status    # 対応状況
            if not row[6].value:
                row[6].value = TODAY  # 完了日
            count += 1
    print(f'機能・仕様要望: {count}件を完了化')

    wb.save(REQ_XLSX)


def update_spec_workbook():
    wb = openpyxl.load_workbook(SPEC_XLSX)

    # --- Revision: v1.6 追加 ---
    ws = wb['Revision']
    last = None
    for row in ws.iter_rows(min_row=1):
        if row[0].value is not None:
            last = row[0].row
    target = last + 1
    ws.cell(row=target, column=1).value = 'v1.6'
    ws.cell(row=target, column=2).value = TODAY
    ws.cell(row=target, column=3).value = (
        '速さランキング追加(scores.avg_answer_time / GET /rankingのspeed / ランキング画面のFastest Answers)、'
        'ローカルデータモード(DATA_SOURCE=local)追加、エラーコード体系({ok, code, message})導入、'
        'TTSの常時SSML化(発音補正・頭切れ対策・キャッシュ版数化)、'
        'イラストあり問題のレイアウト分離、レート制限・helmet・平文パスワード透過移行'
    )
    ws.cell(row=target, column=4).value = '藤岡諒也'
    print(f'Revision: v1.6 を{target}行目に追加')

    # --- Data Specs: scores に avg_answer_time 追加 ---
    ws = wb['Data Specs']
    play_date_row = None
    for row in ws.iter_rows(min_row=1):
        # scoresブロックのplay_date行を探す(C列=カラム名)
        if str(row[2].value or '').strip() == 'play_date':
            play_date_row = row[0].row
    if play_date_row:
        ws.insert_rows(play_date_row + 1)
        r = play_date_row + 1
        ws.cell(row=r, column=3).value = 'avg_answer_time'
        ws.cell(row=r, column=4).value = 'number'
        ws.cell(row=r, column=5).value = '平均回答秒(回答受付開始から正解までの平均。正解なしのプレイは空)'
        ws.cell(row=r, column=6).value = '速さランキングに使用'
        print(f'Data Specs: avg_answer_time を{r}行目に追加')
    else:
        print('WARN: Data Specsのplay_date行が見つからない')

    # --- Function List: ランキング3種化 ---
    ws = wb['Function List']
    for row in ws.iter_rows(min_row=1):
        name = str(row[2].value or '')
        if name == 'ランキングデータ取得':
            row[3].value = '当月scoresを集計し、挑戦回数・平均スコア・平均回答秒(速さ)の上位3名を返す'
            row[5].value = 'month、challenge上位3、accuracy上位3、speed上位3'
            print(f'Function List: ランキングデータ取得({row[0].row}行目)を更新')
        elif name == 'ランキング表示':
            row[3].value = 'ランキングAPIの結果を3パネルで表示'
            row[5].value = '挑戦回数/Best Scores/Fastest AnswersのNo.1〜No.3'
            print(f'Function List: ランキング表示({row[0].row}行目)を更新')

    # --- UI Specs: G005 ---
    ws = wb['UI Specs']
    for row in ws.iter_rows(min_row=1):
        if str(row[0].value or '').strip() == 'G005':
            row[2].value = ('・Period表示 \n・Number of try 上位3 \n・Best Scores 上位3 \n'
                            '・Fastest Answers 上位3(平均回答秒) \n・Backボタン')
            print(f'UI Specs: G005({row[0].row}行目)を更新')

    # --- Tech Stack: バックエンド設定にDATA_SOURCE追記 ---
    ws = wb['Tech Stack']
    for row in ws.iter_rows(min_row=1):
        if str(row[0].value or '').strip() == 'バックエンド設定':
            row[1].value = ('SHEET_ID / GOOGLE_KEYFILE / GOOGLE_CREDENTIALS_JSON / JWT_SECRET / '
                            'FRONTEND_URL / REDIS_* / DATA_SOURCE')
            row[2].value = ('Google認証、CORS、JWT、Redis接続を環境変数で管理。'
                            'DATA_SOURCE=localでbackend/data/*.jsonを使用したローカル実行(Google認証不要)')
            print(f'Tech Stack: バックエンド設定({row[0].row}行目)を更新')

    wb.save(SPEC_XLSX)


def main():
    update_requests_workbook()
    update_spec_workbook()
    print('OK: ワークブック更新完了')
    return 0


if __name__ == '__main__':
    sys.exit(main())
